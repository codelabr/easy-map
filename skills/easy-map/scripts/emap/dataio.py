"""Files in, files out: dependencies, workbooks, shapefiles, run folders."""

from __future__ import annotations

import os

import csv
import json
import re
import shutil
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Sequence

from . import detect, messages as msg

EXCEL_SUFFIXES = (".xlsx", ".xlsm", ".xls")

#: A pasted table has to land somewhere, and a CSV is the only honest place: it
#: keeps one row per line and one delimiter, which is exactly what a paste is.
#: Accepted from ``input/`` as well, because a user who exports CSV from their
#: system should not have to open Excel first just to be allowed in.
TEXT_SUFFIXES = (".csv", ".tsv", ".txt")
TABLE_SUFFIXES = EXCEL_SUFFIXES + TEXT_SUFFIXES

#: the sheet name reported for a file that has no sheets
SINGLE_SHEET = "(csv)"

#: Above this, a workbook is not re-opened to look for merged cells. Reading
#: merges needs openpyxl's full mode, measured at roughly 11 seconds per
#: megabyte — 67 s on the 6 MB PEPFAR export against 0.02 s read-only. Two
#: megabytes puts the worst case near twenty seconds, and reports that use
#: merged cells are far smaller than that: the fixture here is 8 KB. A sheet
#: that trips the ceiling is reported, not silently read the plain way.
MERGE_SCAN_MAX_BYTES = 2_000_000
#: Boundary formats a tier folder may hold. ``.json`` is here because a GeoJSON
#: does not have to be called ``.geojson`` — a Canadian download to hand is
#: ``ca.json`` — and refusing it would be refusing the format on a technicality.
BOUNDARY_SUFFIXES = (".shp", ".geojson", ".json", ".kml")

#: Order of preference when a folder somehow holds more than one candidate that
#: differ only by format. A shapefile keeps its column types, so it wins.
_FORMAT_ORDER = {suffix: i for i, suffix in enumerate(BOUNDARY_SUFFIXES)}

#: A shapefile is not a file. Without these two the geometry or the attribute
#: table is simply absent, and the reader's own message names neither.
_SHAPEFILE_SIDECARS = (".shx", ".dbf")

@dataclass
class Deps:
    pd: Any
    gpd: Any = None
    plt: Any = None
    matplotlib: Any = None


def load(require_geo: bool = True, require_plot: bool = False) -> Deps:
    try:
        import pandas as pd
    except ImportError as exc:  # pragma: no cover
        raise SystemExit(msg.text("loi.thiếu-thư-viện", library="pandas")) from exc

    deps = Deps(pd=pd)
    if not require_geo:
        # Not required is not the same as not wanted. ``list`` runs without
        # geopandas on purpose — it is the command someone reaches for when
        # something is broken — but where the library is there, it can say a
        # great deal more, and refusing to look would be a strange way to keep
        # a promise about not requiring it.
        try:
            import geopandas as gpd

            deps.gpd = gpd
        except ImportError:
            pass
    if require_geo:
        try:
            import geopandas as gpd
        except ImportError as exc:
            raise SystemExit(msg.text("loi.thiếu-thư-viện", library="geopandas")) from exc
        deps.gpd = gpd
    if require_plot:
        try:
            import matplotlib
            matplotlib.use("Agg")
            import matplotlib.pyplot as plt
        except ImportError as exc:
            raise SystemExit(msg.text("loi.thiếu-thư-viện", library="matplotlib")) from exc
        deps.matplotlib, deps.plt = matplotlib, plt
    return deps


def project_path(project_root: Path, raw: str | None) -> Path | None:
    if not raw:
        return None
    p = Path(raw)
    return p if p.is_absolute() else (Path(project_root) / p)


def find_excel_files(project_root: Path) -> list[Path]:
    folder = Path(project_root) / "input"
    if not folder.exists():
        return []
    return sorted(p for p in folder.iterdir()
                  if p.suffix.lower() in TABLE_SUFFIXES and not p.name.startswith("~$"))


def is_text_table(path: Path) -> bool:
    return Path(path).suffix.lower() in TEXT_SUFFIXES


#: characters a file name cannot carry on Windows; everything else, including
#: Vietnamese diacritics and spaces, is left alone so the user still recognises
#: their own file in the list
_UNSAFE_NAME = re.compile(r'[<>:"/\\|?*\x00-\x1f]')


def safe_name(raw: str) -> str:
    cleaned = _UNSAFE_NAME.sub("_", str(raw)).strip().strip(".")
    return cleaned or "du-lieu"


def _digest(path: Path) -> str:
    import hashlib

    h = hashlib.sha256()
    with open(path, "rb") as handle:
        for chunk in iter(lambda: handle.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def adopt_file(project_root: Path, source: Path) -> dict[str, Any]:
    """Copy a file the user handed over into ``input/``.

    Most people do not put a workbook in a folder — they attach it to the chat.
    That file lives outside the project and does not survive the session, so it
    has to be brought in before anything reads it a second time; a profile and a
    render are two separate reads, minutes apart.

    Never overwrites. Same name **and** same bytes is the same file: say so and
    reuse it, because someone re-attaching a workbook after a failed run should
    not be left choosing between two identical copies. Same name, different
    bytes gets a numbered suffix — the older file may be what an earlier map in
    this very conversation was drawn from.
    """
    source = Path(source)
    if not source.exists() or not source.is_file():
        raise SystemExit(msg.text("loi.không-tìm-thấy-tệp", path=source))
    if source.suffix.lower() not in TABLE_SUFFIXES:
        raise SystemExit(msg.text(
            "loi.định-dạng-không-đọc-được",
            suffix=source.suffix or msg.text("loi.không-có-đuôi"),
            accepted=", ".join(TABLE_SUFFIXES)))

    folder = Path(project_root) / "input"
    folder.mkdir(parents=True, exist_ok=True)
    target = folder / safe_name(source.name)

    if target.exists():
        if target.stat().st_size == source.stat().st_size and \
                _digest(target) == _digest(source):
            return {"tệp": f"input/{target.name}", "trạng_thái": "đã_có_sẵn",
                    "ghi_chú": msg.text("doc.tệp-trùng")}
        stem, suffix = target.stem, target.suffix
        index = 2
        while (folder / f"{stem}_{index:02d}{suffix}").exists():
            index += 1
        target = folder / f"{stem}_{index:02d}{suffix}"

    shutil.copy2(source, target)
    return {"tệp": f"input/{target.name}", "trạng_thái": "đã_chép",
            "tên_gốc": source.name, "dung_lượng_byte": target.stat().st_size}


#: Where the boundary layers live, when they do not live in the project.
#: Set by the installer so a globally installed skill can draw from any working
#: folder: the boundaries are one shared 135 MB set, while the project root is
#: wherever this particular piece of work keeps its input and output.
SHAPEFILE_ENV = "EASY_MAP_SHAPEFILES"


def shapefile_root(project_root: Path, override: str | None = None) -> Path:
    """Explicit flag first, then the environment, then inside the project."""
    if override:
        return Path(override).expanduser()
    from_env = os.environ.get(SHAPEFILE_ENV)
    if from_env:
        return Path(from_env).expanduser()
    return Path(project_root) / "shapefiles"


#: The layout the project shipped with, and what each folder becomes. Vietnam
#: is named here because these two folder names only ever held Vietnam.
LEGACY_TIERS = {"provinces": ("viet-nam", "province"),
                "communes": ("viet-nam", "commune")}

#: The two roles a tier can play. Which folder plays which is decided by
#: counting features, never by reading the folder's name: ``provinces`` sorts
#: before ``communes`` in English and ``region`` before ``district``, but
#: ``comuna`` sorts before ``judet`` and that would be backwards.
COARSE, FINE = "province", "commune"


def migrate_legacy_layout(root: Path) -> list[dict[str, Any]]:
    """Move ``provinces/`` and ``communes/`` under ``viet-nam/``.

    A rename, not a copy: the Vietnamese boundaries are 135 MB and copying them
    to reshape a folder would be a strange thing to do to somebody's disk.

    Refuses to overwrite. If ``viet-nam/province/`` already holds something,
    the old folder is left exactly where it is and reported, because the two
    could be different data and only the user knows which is wanted.
    """
    moved = []
    for legacy, (country, tier) in LEGACY_TIERS.items():
        source = root / legacy
        if not source.is_dir():
            continue
        target = root / country / tier
        if target.exists():
            moved.append({"từ": legacy, "sang": f"{country}/{tier}",
                          "trạng_thái": "bỏ_qua_vì_đích_đã_có"})
            continue
        target.parent.mkdir(parents=True, exist_ok=True)
        source.rename(target)
        moved.append({"từ": legacy, "sang": f"{country}/{tier}",
                      "trạng_thái": "đã_chuyển"})
    return moved


def _has_boundary(folder: Path) -> bool:
    return folder.is_dir() and any(
        f.is_file() and f.suffix.lower() in BOUNDARY_SUFFIXES
        for f in folder.iterdir())


def countries(root: Path) -> list[str]:
    """Country folder names, in the order they should be offered."""
    if not root.is_dir():
        return []
    return sorted(d.name for d in root.iterdir()
                  if d.is_dir() and not d.name.startswith(".")
                  and any(_has_boundary(t) for t in d.iterdir() if t.is_dir()))


def resolve_country(root: Path, requested: str | None = None) -> str:
    if requested:
        if not _has_boundary_country(root, requested):
            raise SystemExit(msg.text("loi.không-có-quốc-gia", country=requested,
                                      available=", ".join(countries(root)) or "-"))
        return requested
    found = countries(root)
    if not found:
        raise SystemExit(msg.text("loi.thiếu-thư-mục-shapefile", folder=root))
    if len(found) > 1:
        raise SystemExit(msg.text("loi.nhiều-quốc-gia",
                                  available=", ".join(found)))
    return found[0]


def _has_boundary_country(root: Path, name: str) -> bool:
    folder = root / name
    return folder.is_dir() and any(_has_boundary(t) for t in folder.iterdir()
                                   if t.is_dir())


def _feature_count(path: Path) -> int | None:
    """How many features, without reading the geometry, or None.

    The commune layer is 115 MB and 3,321 features; opening it to count them on
    every command would make the tier order cost more than the map.

    None rather than an exception when the file cannot be opened, because this
    runs while *listing* what is available. A boundary file that cannot be read
    is a real error, but the place to raise it is where the map is drawn, with
    a message about that file — not here, where it would take down the one
    command whose job is to say what is present.
    """
    try:
        import pyogrio

        return int(pyogrio.read_info(str(path))["features"])
    except Exception:                      # pragma: no cover - depends on driver
        try:
            import geopandas

            return len(geopandas.read_file(path))
        except Exception:
            return None


def tiers(root: Path, country: str) -> list[dict[str, Any]]:
    """The country's tiers, coarsest first, with the role each one plays.

    Order comes from the feature count. A country with one tier gets the coarse
    role and nothing else, which is the case the old two-folder layout could
    not express at all: a boundary set that only goes down to states could not
    be loaded, even to draw the states.
    """
    folder = root / country
    if not folder.is_dir():
        raise SystemExit(msg.text("loi.không-có-quốc-gia", country=country,
                                  available=", ".join(countries(root)) or "-"))
    found = []
    for tier in sorted(folder.iterdir()):
        if not _has_boundary(tier):
            continue
        path = _one_dataset(tier)
        found.append({"thư_mục": tier.name, "tệp": path.name,
                      "số_đơn_vị": _feature_count(path), "__path": path})
    # Unreadable files sort last and keep their order, so one bad file cannot
    # silently promote itself to the coarse tier and become the country.
    found.sort(key=lambda t: (t["số_đơn_vị"] is None, t["số_đơn_vị"] or 0))

    # A GADM download holds four files and the first of them is the outline of
    # the whole country: one feature, and therefore the smallest tier of the
    # set. Ranked by size it would take the coarsest role and every province
    # would be matched against a single shape called "Vietnam". The file is
    # opened only when the count is one, which costs one feature to read.
    for entry in found:
        entry["là_đường_viền_quốc_gia"] = False
        if entry["số_đơn_vị"] == 1:
            try:
                import geopandas

                frame = geopandas.read_file(entry["__path"])
            except Exception:               # pragma: no cover - depends on driver
                continue
            entry["là_đường_viền_quốc_gia"] = bool(
                detect.identify(frame).get("là_đường_viền_quốc_gia"))

    ranked = [t for t in found if not t["là_đường_viền_quốc_gia"]]
    for entry in found:
        entry["vai_trò"] = None
    for role, entry in zip((COARSE, FINE), ranked):
        entry["vai_trò"] = role
    return found


def find_boundaries(project_root: Path, admin_level: str,
                    override: str | None = None,
                    country: str | None = None) -> Path:
    """The one boundary file for a tier, whatever format and country.

    ``admin_level`` accepts either the role — ``province`` or ``commune`` — or
    the tier folder's own name, so a United States boundary set can be asked
    for as ``state`` and a Vietnamese one as ``province``, and both work.
    """
    root = shapefile_root(project_root, override)
    migrate_legacy_layout(root)
    name = resolve_country(root, country)
    available = tiers(root, name)
    if not available:
        raise SystemExit(msg.text("loi.thiếu-thư-mục-shapefile", folder=root / name))

    for entry in available:
        if entry.get("là_đường_viền_quốc_gia"):
            continue
        if admin_level in (entry["thư_mục"], entry.get("vai_trò")):
            return entry["__path"]
    raise SystemExit(msg.text(
        "loi.không-có-tầng", level=admin_level, country=name,
        available=", ".join(f"{t['thư_mục']} ({t['số_đơn_vị']})" for t in available)))


def resolve_tier(project_root: Path, admin_level: str,
                 override: str | None = None,
                 country: str | None = None) -> dict[str, Any]:
    """The tier entry a request names, whether by role or by folder name.

    Everything downstream asks ``admin_level == "commune"`` to decide how to
    behave, so a request for a folder called ``district`` has to become the
    role ``commune`` here and not somewhere later. Leaving the folder name in
    circulation would send a fine tier down every coarse-tier branch, quietly.
    """
    root = shapefile_root(project_root, override)
    migrate_legacy_layout(root)
    name = resolve_country(root, country)
    for entry in tiers(root, name):
        if entry.get("là_đường_viền_quốc_gia"):
            continue
        if admin_level in (entry["thư_mục"], entry.get("vai_trò")):
            return {**entry, "quốc_gia": name}
    raise SystemExit(msg.text(
        "loi.không-có-tầng", level=admin_level, country=name,
        available=", ".join(f"{t['thư_mục']} ({t['số_đơn_vị']})"
                            for t in tiers(root, name))))


def _one_dataset(folder: Path) -> Path:
    """The single boundary dataset in a tier folder.

    A tier folder holds exactly one dataset. Two datasets in one folder is the
    kind of mistake that draws the wrong map without anyone noticing, so it is
    refused by name rather than resolved by sorting.
    """
    if not folder.exists():
        raise SystemExit(msg.text("loi.thiếu-thư-mục-shapefile", folder=folder))

    found = sorted((p for p in folder.iterdir()
                    if p.is_file() and p.suffix.lower() in BOUNDARY_SUFFIXES),
                   key=lambda p: (_FORMAT_ORDER[p.suffix.lower()], p.name))
    if not found:
        raise SystemExit(msg.text("loi.không-có-ranh-giới", folder=folder,
                                  accepted=", ".join(BOUNDARY_SUFFIXES)))
    if len(found) > 1 and len({p.stem for p in found}) > 1:
        raise SystemExit(msg.text("loi.nhiều-tệp-ranh-giới", folder=folder,
                                  files=", ".join(p.name for p in found)))

    chosen = found[0]
    if chosen.suffix.lower() == ".shp":
        missing = [s for s in _SHAPEFILE_SIDECARS
                   if not chosen.with_suffix(s).exists()]
        if missing:
            raise SystemExit(msg.text("loi.thiếu-tệp-đi-kèm", path=chosen.name,
                                      missing=", ".join(missing)))
    return chosen


def read_sheets(deps: Deps, excel: Path) -> list[str]:
    if is_text_table(excel):
        return [SINGLE_SHEET]
    book = deps.pd.ExcelFile(excel)
    try:
        return list(book.sheet_names)
    finally:
        book.close()


def _text_dialect(path: Path) -> dict[str, Any]:
    """How to read a delimited text file, decided from the file itself.

    A pasted table arrives however the source happened to write it: tabs from a
    spreadsheet, semicolons from a Vietnamese-locale export, commas from a
    normal CSV. Guessing wrong does not raise — it produces one column holding
    the whole line, which then looks like a sheet with no place-name column.
    Encoding gets the same treatment: files from Windows tools are frequently
    UTF-8 with a BOM or cp1258, and a wrong guess turns 'Hà Nội' into mojibake
    that no longer matches the shapefile.
    """
    raw = Path(path).read_bytes()[:64_000]
    encoding = "utf-8-sig"
    for candidate in ("utf-8-sig", "utf-8", "cp1258", "latin-1"):
        try:
            text = raw.decode(candidate)
        except UnicodeDecodeError:
            continue
        encoding = candidate
        break
    else:                                     # pragma: no cover - latin-1 never fails
        text = raw.decode("latin-1")

    sample = "\n".join(text.splitlines()[:20])
    try:
        sep = csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except csv.Error:
        # Sniffer gives up on a single-column file; count instead, and prefer
        # the tab because that is what a spreadsheet paste produces
        counts = {c: sample.count(c) for c in "\t;,|"}
        sep = max(counts, key=counts.get) if any(counts.values()) else ","
    return {"encoding": encoding, "sep": sep}


def read_text_table(deps: Deps, path: Path, notes: list[dict[str, Any]] | None = None):
    dialect = _text_dialect(path)
    df = deps.pd.read_csv(path, dtype=str, keep_default_na=False,
                          na_values=[""], **dialect)
    if notes is not None:
        notes.append({
            "việc": "đọc_văn_bản_phân_cách",
            "chi_tiết": msg.text("doc.bảng-dán", encoding=dialect["encoding"],
                                 delimiter=repr(dialect["sep"])),
        })
    return df


def read_table(deps: Deps, excel: Path, sheet: str | None,
               notes: list[dict[str, Any]] | None = None):
    """The sheet as a tidy frame, with a real workbook's habits accounted for.

    Two of those habits used to go straight through: a header sitting below a
    title block or a set of pivot filters, and numbers typed with thousands
    separators so the column arrives as text. Neither raises anything — the
    first produces columns called ``Unnamed: 3`` and the second produces a
    column nobody can map.

    Anything adjusted is appended to ``notes`` rather than done quietly.
    """
    from . import tabular

    if is_text_table(excel):
        df = read_text_table(deps, excel, notes)
        start = 0
    else:
        peek = deps.pd.read_excel(excel, sheet_name=sheet or 0, header=None,
                                  nrows=tabular.MAX_HEADER_SCAN + 5)
        start = tabular.header_row(peek.values.tolist())
        df = deps.pd.read_excel(excel, sheet_name=sheet or 0,
                                header=start if start else 0)
        df.columns = [str(c).strip() for c in df.columns]
        merged = _read_merged(deps, excel, sheet, df, notes)
        if merged is not None:
            df, start = merged, 0
    df.columns = [str(c).strip() for c in df.columns]
    if start and notes is not None:
        notes.append({
            "việc": "bỏ_qua_dòng_đầu",
            "chi_tiết": msg.text("doc.dòng-tiêu-đề", row=start + 1, skipped=start),
        })

    for column in df.columns:
        # ask what the column *is*, not what dtype object it happens to carry:
        # a CSV read with dtype=str arrives as the "str" dtype on current pandas,
        # which is not `object`, so a `!= object` test skipped every column and
        # left a table of numbers looking like a table of categories
        if deps.pd.api.types.is_numeric_dtype(df[column]):
            continue
        converted = tabular.coerce_column(df[column].tolist())
        if converted is None:
            continue
        values, note = converted
        df[column] = deps.pd.to_numeric(deps.pd.Series(values, index=df.index),
                                        errors="coerce")
        if notes is not None:
            notes.append({"việc": "đổi_chữ_thành_số", "cột": column, **note})
    return df


def _read_merged(deps: Deps, excel: Path, sheet: str | None, plain, notes):
    """Re-read the sheet honouring merged cells, or None if that is not needed.

    An agency report writes a province once and merges it down over its own
    rows, and groups two columns under one heading on the row above. Read
    plainly, the first habit blanks out most of the place names and the second
    turns the second tier of the header — "Nam", "Nữ" — into a row of data.
    Neither raises.
    """
    from . import tabular

    first = plain.iloc[:, 0].tolist() if plain.shape[1] else []
    if not tabular.looks_merged(list(plain.columns), first):
        return None
    if excel.stat().st_size > MERGE_SCAN_MAX_BYTES:
        if notes is not None:
            notes.append({
                "việc": "bỏ_qua_dò_ô_gộp",
                "chi_tiết": msg.text("doc.ô-gộp-tệp-lớn",
                                     limit=MERGE_SCAN_MAX_BYTES // 1_000_000),
            })
        return None

    from openpyxl import load_workbook

    book = load_workbook(excel, data_only=True)
    try:
        ws = book[sheet] if sheet else book.worksheets[0]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        ranges = [(r.min_row - 1, r.min_col - 1, r.max_row - 1, r.max_col - 1)
                  for r in ws.merged_cells.ranges]
    finally:
        book.close()
    if not ranges:
        return None

    rows = tabular.fill_merged(rows, ranges)
    start = tabular.header_top(tabular.header_row(rows), ranges)
    depth = tabular.header_depth(start, ranges)
    names = tabular.join_header(rows, start, depth)
    body = [r for r in rows[start + depth:]
            if any(not tabular.is_blank(c) for c in r)]
    width = len(names)
    body = [list(r[:width]) + [None] * (width - len(r)) for r in body]

    if notes is not None:
        notes.append({
            "việc": "đọc_ô_gộp",
            "số_vùng_gộp": len(ranges),
            "số_tầng_tiêu_đề": depth,
            "chi_tiết": msg.text(
                "doc.ô-gộp", regions=len(ranges),
                header=(msg.text("doc.ô-gộp-tiêu-đề", levels=depth,
                                 example=names[3] if len(names) > 3 else names[-1])
                        if depth > 1 else "") + ".",),
        })
    return deps.pd.DataFrame(body, columns=names)


def read_data_dictionary(deps: Deps, excel: Path, sheets: Sequence[str]) -> dict[str, str] | None:
    """Use a 'Từ điển dữ liệu' sheet as the authoritative meaning of columns."""
    target = None
    for name in sheets:
        low = name.lower()
        if "từ điển" in low or "tu dien" in low or "dictionary" in low or "data dict" in low:
            target = name
            break
    if target is None:
        return None
    try:
        df = read_table(deps, excel, target)
    except Exception:
        return None
    if df.shape[1] < 2:
        return None
    key_col, desc_col = df.columns[0], df.columns[1]
    out: dict[str, str] = {}
    for _, row in df.iterrows():
        key = str(row[key_col]).strip()
        val = str(row[desc_col]).strip()
        if key and key.lower() != "nan":
            out[key] = val
    return out or None


#: The signature of UTF-8 text decoded as Latin-1: every multi-byte character
#: turns into a run starting with one of these. Used only to skip work — the
#: decision is the round trip below, never the presence of a character.
_MOJIBAKE_HINTS = ("Ã", "Â", "Ä", "Å", "á", "â", "ã", "ð", "ï")


def demojibake(text: str) -> str | None:
    """The same text read the way it was written, or None if it already was.

    A DBF does not record its own encoding. The convention is a companion
    ``.cpg`` file naming the codepage; without one the reader falls back to
    Latin-1, and UTF-8 bytes come back as mojibake — ``é`` as ``Ã©``.

    The test is a round trip, not a list of suspicious characters: encode back
    to the bytes Latin-1 would have produced, then decode them as UTF-8. That
    only succeeds when the bytes really were UTF-8, because UTF-8 has a shape a
    random Latin-1 string does not accidentally have. A genuine Latin-1 name
    fails the round trip and is left alone.
    """
    if not any(hint in text for hint in _MOJIBAKE_HINTS):
        return None
    try:
        fixed = text.encode("latin-1").decode("utf-8")
    except (UnicodeEncodeError, UnicodeDecodeError):
        return None
    return fixed if fixed != text else None


def _encoding_repair(gdf, path: Path) -> dict[str, Any] | None:
    """Whether this frame is a mis-decoded shapefile, and what it costs.

    Only shapefiles can have the problem, and only those with no ``.cpg``. The
    evidence has to be unanimous: every value that looks mis-decoded must
    repair cleanly. One value that does not is enough to leave the file alone,
    because a half-repaired attribute table is worse than an unrepaired one.
    """
    if path.suffix.lower() != ".shp" or path.with_suffix(".cpg").exists():
        return None

    repaired, sample = 0, None
    for column in gdf.columns:
        # No dtype filter. pandas hands text back as ``str`` on some versions
        # and ``object`` on others, and a filter written for one of them skips
        # every column on the other — silently, because a file with nothing to
        # repair and a file that was never looked at are indistinguishable from
        # here. ``isinstance`` below is the only test that needs to be right.
        if column == "geometry":
            continue
        for value in gdf[column].dropna():
            if not isinstance(value, str):
                continue
            fixed = demojibake(value)
            if fixed is None:
                continue
            repaired += 1
            if sample is None:
                sample = (value, fixed)
    if not repaired:
        return None
    # Named by its folder as well as its file, because the two tiers are
    # repaired separately and two lines reading "ca.shp" look like the same
    # thing reported twice when they are two files.
    return {"tệp": f"{path.parent.name}/{path.name}", "số_giá_trị": repaired,
            "ví_dụ": {"đọc_sai": sample[0], "đúng": sample[1]},
            "cách_sửa": f"{path.parent.name}/{path.stem}.cpg"}


def load_shapes(deps: Deps, project_root: Path, admin_level: str,
                override: str | None = None, notes: list | None = None,
                country: str | None = None):
    """The boundary frame for one tier, in whatever format it was supplied.

    Reading is the same call for all three formats — the driver is chosen from
    the suffix — so the only thing this has to do beyond reading is repair a
    shapefile whose codepage was never recorded. A repair is never silent: it
    is appended to ``notes`` so the caller can say what it changed.
    """
    path = find_boundaries(project_root, admin_level, override, country)
    gdf = deps.gpd.read_file(path)

    repair = _encoding_repair(gdf, path)
    if repair is not None:
        gdf = deps.gpd.read_file(path, encoding="utf-8")
        if notes is not None:
            notes.append(repair)

    gdf = gdf.reset_index(drop=True)
    gdf["__shape_id"] = gdf.index
    return gdf


def shape_fields(gdf, admin_level: str) -> dict[str, str | None]:
    """Where the names are, in the two slots the rest of the engine expects.

    ``province`` is the coarse tier's name column, or — on a fine-tier frame —
    the column naming its parent. ``commune`` is the fine tier's own name
    column. The two slots are older than the multi-country work and the names
    on them are Vietnamese, but they are roles rather than places, and the
    engine branches on them in a dozen spots; renaming them would be a large
    change with no reader benefit.

    What is new is that neither is looked up in a fixed list any more. The list
    held ``ten_tinh`` and four spellings around it, which is right for exactly
    one dataset in the world.
    """
    reading = detect.identify(gdf)
    if reading.get("là_đường_viền_quốc_gia"):
        raise SystemExit(msg.text("loi.là-đường-viền-quốc-gia",
                                  evidence=reading["bằng_chứng"]))

    name = reading.get("cột_tên")
    if name is None:
        raise SystemExit(msg.text("loi.không-tìm-được-cột-tên", level=admin_level,
                                  evidence=reading["bằng_chứng"]))
    # Two keys and no more. The reading that produced them, with its evidence,
    # belongs in the country profile where an agent can read it once; carried
    # back here it would ride along inside a dict a dozen callers index by
    # name and none of them expect to be a report.
    if admin_level == FINE:
        return {"province": reading.get("cột_cha"), "commune": name}
    return {"province": name, "commune": None}


#: Latitude beyond which a standard parallel stops meaning anything.
_LAT_LIMIT = 89.0

#: Smallest latitude span the standard parallels are spread over. A country
#: only a fraction of a degree tall would otherwise get two parallels almost on
#: top of each other, which is where Albers degenerates.
_MIN_LAT_SPAN = 1.0


def thematic_crs(gdf) -> str:
    """An equal-area projection sized for whatever country this frame holds.

    Equal-area is not the choice being made here — the map compares quantities
    across units, so areas have to be comparable and that settles it. What is
    being chosen is where to centre the projection, and getting that wrong is
    not a matter of taste: projecting the United States into a projection
    centred on Vietnam turned two valid state outlines into self-intersecting
    ones and killed the run inside GEOS, with a message that named neither the
    projection nor the country.

    **The central meridian is a mean of angles, not of numbers.** Averaging
    longitudes arithmetically puts the United States in the Gulf of Guinea,
    because Alaska's Aleutians reach past the antimeridian and −179 and 179 are
    a quarter of a degree apart, not 358. Summing unit vectors and taking the
    direction of the total is the ordinary way to average a direction, and it
    makes the antimeridian a non-event rather than a special case.

    **Weighted by area**, so that a scatter of small offshore fragments does not
    drag the centre out to sea. This replaces the meridian-splitting rule that
    was written for Vietnam: there is no meridian that separates the mainland
    United States from Alaska, and a rule that only works where a country
    happens to have its islands on one side is not a rule.

    Measured on 92 countries plus Vietnam at both tiers, the United States,
    Canada and the fixture: every one projects and unions without error, and
    the only invalid geometry produced was invalid in the source file already.
    Against the values worked out by hand it lands within a few hundredths of a
    degree for Vietnam and the fixture, and 2.5 degrees for Canada.

    Where it is *not* good: a country with a large detached territory is pulled
    toward it — the United States comes out at −112.4 where an atlas would use
    −96, because Alaska is a fifth of the country's area. The map is usable and
    equal-area; the shapes of the lower 48 are sheared. The proper answer is to
    frame the main body and carry the rest in an inset, which is what Vietnam
    already does by a hard-coded meridian and what a later round has to
    generalise. Until then this is a known cost, not an unnoticed one.
    """
    import math
    import warnings

    parts = gdf.explode(index_parts=False, ignore_index=True)
    centres = parts.geometry.representative_point()
    with warnings.catch_warnings():
        # geopandas warns that an area in degrees is not an area, and it is
        # right — but nothing here is an area. These are relative weights, and
        # any monotone measure of size serves. Letting the warning through
        # would put a sentence about incorrect results on the console of a run
        # that is correct, which is worse than useless.
        warnings.filterwarnings("ignore", message=".*geographic CRS.*")
        weights = parts.geometry.area
    if float(weights.sum()) <= 0:                      # points, or degenerate
        weights = weights * 0 + 1

    east = float((weights * centres.x.map(math.radians).map(math.cos)).sum())
    north = float((weights * centres.x.map(math.radians).map(math.sin)).sum())
    lon_0 = math.degrees(math.atan2(north, east))

    bounds = gdf.total_bounds
    low, high = float(bounds[1]), float(bounds[3])
    if high - low < _MIN_LAT_SPAN:
        pad = (_MIN_LAT_SPAN - (high - low)) / 2
        low, high = low - pad, high + pad
    span = high - low

    # The two-sixths rule: standard parallels a sixth of the way in from each
    # edge. It is the ordinary cartographic choice for a conic, and it is what
    # the hand-picked Vietnamese parallels turn out to have been.
    def clamp(value: float) -> float:
        return max(-_LAT_LIMIT, min(_LAT_LIMIT, value))

    return (f"+proj=aea +lat_1={clamp(low + span / 6):.4f} "
            f"+lat_2={clamp(high - span / 6):.4f} "
            f"+lat_0={clamp((low + high) / 2):.4f} +lon_0={lon_0:.4f} "
            f"+datum=WGS84 +units=m +no_defs")


#: One projection per country per run, resolved from the coarsest tier and
#: reused. Inferring it separately for each frame would centre the map and its
#: locator on very slightly different meridians — 106.4149 against 106.4126 for
#: Vietnam — which is small enough to look like nothing and wrong all the same.
_CRS_CACHE: dict[str, str] = {}


def run_thematic_crs(deps: Deps, project_root: Path,
                     override: str | None = None,
                     country: str | None = None) -> str:
    """The projection for this run, taken from the province tier.

    The province tier rather than whichever tier is being drawn, because a map
    of the communes of one province must sit in the same projection as the
    national locator beside it.
    """
    root = shapefile_root(project_root, override)
    migrate_legacy_layout(root)
    name = resolve_country(root, country)
    key = f"{root}|{name}"
    if key not in _CRS_CACHE:
        # Read, not recomputed. The profile already decided this, with its
        # evidence beside it; deriving it a second time here would be a second
        # chance to answer differently.
        _CRS_CACHE[key] = read_country(deps, root, name)["phép_chiếu"]["crs"]
    return _CRS_CACHE[key]


#: Where the reading of each country is kept, beside the boundaries it
#: describes rather than inside any one project, because the boundaries are one
#: shared set and the reading belongs to them.
PROFILE = "ho_so_quoc_gia.json"

#: Raised whenever this file learns a new field. The cache is keyed on the
#: boundary files, which is right for "the data changed" and useless for "the
#: engine changed": a profile written before ``khung_phụ`` existed is valid by
#: that key and answers None to a question it was never asked, so Vietnam loses
#: its inset on every machine that already had a profile and nothing says why.
#: Bump this in the same commit as any new field.
PROFILE_VERSION = 2


def _profile_key(root: Path, country: str) -> list[str]:
    """What the profile was read from, so a changed folder invalidates it.

    Names and sizes, not contents: hashing 135 MB on every command to notice a
    file nobody touched would cost more than the reading it protects.
    """
    key = []
    for tier in sorted((root / country).iterdir()):
        if not tier.is_dir():
            continue
        for f in sorted(tier.iterdir()):
            if f.is_file():
                key.append(f"{tier.name}/{f.name}:{f.stat().st_size}")
    return key


def read_country(deps: Deps, root: Path, country: str,
                 rebuild: bool = False) -> dict[str, Any]:
    """Everything inferred about one country, computed once and kept.

    Every command needs the same handful of answers — which column holds the
    names, which tier is which, where to centre the projection — and each is a
    decision. Made separately in five places they are five chances to disagree
    by a hundredth of a degree or by one column, which is the kind of
    difference nobody sees until a map is wrong.

    Each field carries where it came from and what the evidence was. The
    evidence is not decoration: an agent reading "34/34 of the finer tier's
    parents are known names" can keep quiet, and one reading "26/34" has to
    ask, and nothing but the evidence separates those two.
    """
    store = root / PROFILE
    try:
        saved = json.loads(store.read_text(encoding="utf-8"))
    except (OSError, ValueError):
        saved = {}
    from . import insets

    key = _profile_key(root, country)
    kept = saved.get(country, {})
    if not rebuild and kept.get("__nguồn") == key \
            and kept.get("__phiên_bản") == PROFILE_VERSION:
        # The declaration is the one field a person edits by hand, so it is read
        # back on every command instead of being cached with the machine's
        # reading. The cache is keyed on the boundary files, and editing this
        # file changes none of them: without this, somebody writes a meridian,
        # runs the command, and gets the map they had before.
        fresh = insets.declaration(kept.get("tên_quốc_gia") or country,
                                   kept.get("khai_báo"), where=str(store))
        if fresh != kept.get("khung_phụ"):
            kept["khung_phụ"] = fresh
            _write_profile(store, saved)
        return kept

    reading: dict[str, Any] = {"__phiên_bản": PROFILE_VERSION,
                               "__nguồn": key, "tầng": []}
    # What a person wrote about this country is carried across the rebuild
    # untouched. Everything else in this file is the machine's reading and is
    # thrown away the moment a boundary file changes size; a declaration is not
    # a reading, and losing it because somebody replaced a shapefile would undo
    # a decision nobody asked to undo.
    by_hand = kept.get("khai_báo")
    if by_hand:
        reading["khai_báo"] = by_hand
    frames = {}
    for entry in tiers(root, country):
        gdf = deps.gpd.read_file(entry["__path"])
        found = detect.identify(gdf)
        frames[entry["thư_mục"]] = (gdf, found)
        reading["tầng"].append({
            "thư_mục": entry["thư_mục"], "tệp": entry["tệp"],
            "số_đơn_vị": entry["số_đơn_vị"], "vai_trò": entry.get("vai_trò"),
            "cột_tên": found.get("cột_tên"), "cấp": found.get("cấp"),
            "là_đường_viền_quốc_gia": bool(found.get("là_đường_viền_quốc_gia")),
        })

    named = [t for t in reading["tầng"] if not t["là_đường_viền_quốc_gia"]]
    first = frames.get(named[0]["thư_mục"]) if named else None
    if first is not None:
        gdf, found = first
        reading["nhận_diện"] = {k: found[k] for k in ("bộ", "độ_tin_cậy", "bằng_chứng")}
        reading["tên_quốc_gia"] = found.get("quốc_gia")
        crs = thematic_crs(gdf)
        reading["phép_chiếu"] = {
            "crs": crs, "nguồn": "suy diễn từ hình học của tầng thô",
            "bằng_chứng": f"hộp bao {[round(float(v), 4) for v in gdf.total_bounds]}",
        }
        # Grouping the land takes seconds, so it is done here — once, cached —
        # rather than on every map that might want to mention it.
        reading["lãnh_thổ_rời"] = insets.land_masses(gdf.to_crs(crs))
    # Recorded for every country, including the ones with nothing declared,
    # because "no inset here" is worth telling apart from "nobody has said".
    reading["khung_phụ"] = insets.declaration(
        reading.get("tên_quốc_gia") or country, reading.get("khai_báo"),
        where=str(store))
    if len(named) > 1:
        coarse, fine = frames[named[0]["thư_mục"]], frames[named[1]["thư_mục"]]
        reading["cha_con"] = detect.link_tiers(coarse[0], coarse[1], fine[0], fine[1])

    saved[country] = reading
    _write_profile(store, saved)
    return reading


def _write_profile(store: Path, saved: dict[str, Any]) -> None:
    try:
        store.write_text(json.dumps(saved, ensure_ascii=False, indent=1),
                         encoding="utf-8")
    except OSError:            # a read-only boundary folder is not an error
        pass


def to_thematic_crs(gdf, crs: str):
    if gdf.crs is None:
        return gdf
    return gdf.to_crs(crs)


RUN_STAMP = "%Y-%m-%d_%H-%M-%S"

#: where the name of the request currently in progress is kept
OPEN_RUN = "current-run.json"

#: How long an open run keeps accepting work from a command that forgot to name
#: it. A conversation about one map can run long, but not this long without a
#: single file being written; past that, a nameless command is far more likely
#: to belong to a new request than to the old one.
OPEN_RUN_HOURS = 3.0


def new_run_name() -> str:
    return datetime.now().strftime(RUN_STAMP)


def _open_run_path(project_root: Path) -> Path:
    from . import prefs

    return Path(project_root) / prefs.FOLDER / OPEN_RUN


def remember_open_run(project_root: Path, folder: Path) -> None:
    path = _open_run_path(project_root)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps({"run_folder": folder.name,
                                "mở_lúc": datetime.now().isoformat(timespec="seconds")},
                               ensure_ascii=False, indent=2), encoding="utf-8")


def _touched_at(folder: Path) -> float:
    """Newest write anywhere in the folder, not just the folder's own stamp."""
    stamps = [folder.stat().st_mtime]
    try:
        stamps += [child.stat().st_mtime for child in folder.iterdir()]
    except OSError:
        pass
    return max(stamps)


def open_run(project_root: Path, max_age_hours: float = OPEN_RUN_HOURS) -> str | None:
    """The request in progress, or None if there is none worth reusing."""
    path = _open_run_path(project_root)
    if not path.exists():
        return None
    try:
        name = json.loads(path.read_text(encoding="utf-8")).get("run_folder")
    except (json.JSONDecodeError, OSError):
        return None
    if not name:
        return None
    folder = Path(project_root) / "output" / name
    if not folder.is_dir():
        return None
    if (datetime.now().timestamp() - _touched_at(folder)) > max_age_hours * 3600:
        return None
    return str(name)


def create_run_dir(project_root: Path, requested: str | None = None, *,
                   fresh: bool = False) -> Path:
    """One folder per user request; every artefact of that request lands in it.

    A name passed in explicitly is **reused** when it already exists — that is
    how several commands of the same request (profile, then one render per
    language or layout) share a single folder.

    Naming the folder used to be the only thing holding that contract together,
    so a single command that forgot ``--run-folder`` quietly started a second
    folder and left a stray profile behind. Now an unnamed command joins the run
    already open instead, and only ``fresh=True`` — which is what ``start-run``
    passes — deliberately opens a new one.
    """
    base = Path(project_root) / "output"
    base.mkdir(parents=True, exist_ok=True)

    if not requested and not fresh:
        requested = open_run(project_root)

    if requested:
        folder = base / requested
        folder.mkdir(parents=True, exist_ok=True)
        remember_open_run(project_root, folder)
        return folder

    name = new_run_name()
    folder = base / name
    if folder.exists():
        for i in range(2, 100):
            candidate = base / f"{name}_{i:02d}"
            if not candidate.exists():
                folder = candidate
                break
    folder.mkdir(parents=True)
    remember_open_run(project_root, folder)
    return folder


def slugify(value: str, fallback: str = "ban-do") -> str:
    from .matching import deaccent

    text = re.sub(r"[^a-z0-9]+", "-", deaccent(value)).strip("-")
    return text[:70] or fallback


def link(path: Path | str) -> str:
    """A finished ``file://`` address for something the person is meant to open.

    The engine has always returned real paths and left the agent to make a link
    out of one. A real run rewrote such a path into a ``file://`` URL on a
    different drive, under a mount point that does not exist on the machine the
    file is on: it had guessed at a Linux sandbox it was not running in, and the
    link opened nothing.

    Guessing is not the agent's job here. ``as_uri`` knows the platform it is on,
    escapes the spaces and the Vietnamese in a folder name, and is right on
    Windows and on Linux without being told which it is.
    """
    return Path(path).resolve().as_uri()


#: What a person opens. The side-car JSON and the page's rebuild cache are for
#: the machine, and listing them only makes the real answer harder to find.
OPENABLE = (".png", ".svg", ".html", ".mp4", ".gif", ".csv")


def openable(folder: Path) -> list[dict[str, str]]:
    """Everything this request produced that a person would open, addressed.

    Read off the folder rather than assembled from what each branch happens to
    remember writing: a still render, a video and an interactive page each
    return a different shape, and a list built three times is a list that ends
    up missing whichever one was added last.
    """
    if not Path(folder).is_dir():
        return []
    return [{"tên": p.name, "đường_dẫn": str(p), "liên_kết": link(p)}
            for p in sorted(Path(folder).iterdir())
            if p.is_file() and p.suffix.lower() in OPENABLE]


def write_json(path: Path, payload: Any) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str),
                    encoding="utf-8")
    return path


def write_csv(path: Path, records: Sequence[dict[str, Any]]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not records:
        path.write_text("", encoding="utf-8-sig")
        return path
    fields: list[str] = []
    for r in records:
        for k in r:
            if k not in fields:
                fields.append(k)
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields)
        writer.writeheader()
        writer.writerows(records)
    return path


def read_csv(path: Path) -> list[dict[str, Any]]:
    if not Path(path).exists():
        return []
    with Path(path).open("r", encoding="utf-8-sig", newline="") as fh:
        return list(csv.DictReader(fh))
