#!/usr/bin/env python3
"""easy-map — deterministic engine behind the skill.

Commands
--------
``start-run``  open the folder for one user request and print its name
``list``       what is available: workbooks, sheets, shapefiles, remembered choices
``profile``    evidence about a dataset so the agent can reason like an analyst
``render``     draw and save the finished map(s)
``fix-match``  record a manual name match so future runs reuse it

Output contract: one request, one folder. Call ``start-run`` the moment the user
asks for a map — the folder is stamped ``yyyy-mm-dd_hh-mm-ss`` at *that* moment,
not when rendering finally happens — then pass its name as ``--run-folder`` to
every later command of the same request. Nothing is ever written loose in
``output/``.

``start-run`` is the only command that opens a folder. It leaves that run open,
so a later command that forgets ``--run-folder`` joins it instead of starting a
second one; the run closes itself after a few hours with nothing written to it,
so a new request cannot land in an old folder.

The agent does the talking; this script does the deciding-free work: reading,
matching, aggregating, checking and drawing.

Layout of the skill:

    scripts/easy_map.py   this CLI, the only entry point
    scripts/emap/         the engine, imported by the CLI

Development-only files live outside the skill, at the project root:
``tools/generate_complex_demo.py`` rebuilds the test workbook.
"""

from __future__ import annotations

import argparse
import csv
import json
import shlex
import sys
from pathlib import Path
from typing import Any

sys.path.insert(0, str(Path(__file__).resolve().parent))

from emap import (aggregate, animate, classify, confirm, crosswalk, dataio, detect,  # noqa: E402
                  fonts, furniture, guardrails, i18n, insets, interactive, matching,
                  messages,
                  periods as period_utils,
                  layers, longform, prefs, profile as profiling, render,
                  semantics as sem,
                  tabular, webpage, wording)

PROVINCE_SERIES_THRESHOLD = 5


# --------------------------------------------------------------------------
def speak_utf8() -> None:
    """Print in UTF-8 whatever the console was set to.

    Every reply this engine writes carries Vietnamese - place names at the very
    least - and on Windows Python encodes stdout with the machine's legacy
    codepage rather than UTF-8, for a pipe as much as for a console. One
    accented character then ends the command in UnicodeEncodeError before the
    agent has read anything, and the traceback points at ``json.dumps`` rather
    than at the setting that caused it.

    Guarded rather than assumed: a caller may have replaced the stream with one
    that has no ``reconfigure``, which is what happens when tests capture it.
    """
    for stream in (sys.stdout, sys.stderr):
        try:
            stream.reconfigure(encoding="utf-8")
        except (AttributeError, ValueError, OSError):
            pass


def emit(payload: dict[str, Any]) -> None:
    print(json.dumps(payload, ensure_ascii=False, indent=2, default=str))


def _shape_features(gdf, name_field: str,
                    alias_field: str | None = None) -> list[dict[str, Any]]:
    """One entry per unit, with any second spelling the file gives for it.

    GADM writes an unaccented variant in ``VARNAME_1`` — "Ba Ria - Vung Tau"
    beside "Bà Rịa - Vũng Tàu". Indexing it too turns a table typed without
    accents from a fuzzy guess into an exact hit. The map still shows ``name``:
    an alias is a way in, never a caption.
    """
    aliases = (gdf[alias_field] if alias_field and alias_field in gdf.columns
               else [None] * len(gdf))
    out = []
    for name, alias, shape_id in zip(gdf[name_field], aliases, gdf["__shape_id"]):
        entry: dict[str, Any] = {"name": name, "shape_id": int(shape_id)}
        # GADM writes an absent cell as the two-letter string NA
        text = str(alias).strip() if alias is not None else ""
        if text and text != "NA" and text != str(name).strip():
            entry["aliases"] = [text]
        out.append(entry)
    return out


def _province_index(gdf, name_field: str, affixes, alias_field: str | None = None):
    """Province lookup that also answers to the 63 pre-2025 province names."""
    return matching.build_index(
        crosswalk.alias_features(gdf, name_field=name_field,
                                 alias_field=alias_field), affixes)


def _commune_index_by_province(gdf, fields, affixes,
                               alias_field: str | None = None) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for province, group in gdf.groupby(fields["province"]):
        out[str(province)] = matching.build_index(
            _shape_features(group, fields["commune"], alias_field), affixes)
    return out


#: The bar a commune column has to clear. Defined once in ``tabular`` so the
#: survey and this detector cannot disagree about the same sheet.
COMMUNE_NAME_COVERAGE = tabular.COMMUNE_SHARE


def _name_coverage(df, column, keys, affixes, exclude=frozenset()) -> float:
    """Share of the column's **distinct** names that exist at this level.

    Counting rows instead lets a handful of frequently-repeated names carry the
    vote, which is how 70.000 rows of one export decided it was commune data.
    """
    names = {matching.normalize(v, affixes) for v in df[column].tolist()
             if v is not None}
    names.discard("")
    if not names:
        return 0.0
    return sum(1 for k in names if k in keys and k not in exclude) / len(names)


def _fine_tier(deps, root, country, notes=None):
    """The finer of the two tiers, or nothing.

    A boundary set that stops at states is a complete boundary set, and the old
    two-folder layout could not express it: both folders had to be there or the
    lookup failed, so a country with one tier could not be drawn at all — not
    even at the tier it had.
    """
    try:
        shapes = dataio.load_shapes(deps, root, dataio.FINE, notes=notes,
                                    country=country)
    except SystemExit:
        return None, None
    return shapes, dataio.shape_fields(shapes, dataio.FINE)


def _map_text(pairs: list[str] | None) -> dict[str, str]:
    """``KEY=VALUE`` pairs for strings the engine would otherwise generate.

    Refuses a key it does not know rather than accepting it silently: a
    misspelled key would leave the map in the built-in language while the run
    reported that the text had been set.
    """
    out: dict[str, str] = {}
    for item in pairs or []:
        key, sep, value = str(item).partition("=")
        key = key.strip()
        if not sep or not key:
            raise SystemExit(messages.text("error.map-text-bad-format", item=item))
        if key not in i18n.keys():
            raise SystemExit(messages.text("error.map-text-unknown-key", name=key,
                                           known=", ".join(i18n.keys())))
        out[key] = value
    return out


def _detect_admin_level(df, provinces, communes, affixes) -> str:
    p_keys = {matching.normalize(n, affixes) for n in provinces}
    c_keys = {matching.normalize(n, affixes) for n in communes}
    best = max((_name_coverage(df, c, c_keys, affixes, exclude=p_keys)
                for c in df.columns), default=0.0)
    return "commune" if best >= COMMUNE_NAME_COVERAGE else "province"


# --------------------------------------------------------------------------
def command_start_run(args: argparse.Namespace) -> None:
    """Stamp the folder at the start of the request, before any questions."""
    root = Path(args.project_root).resolve()
    # the one command that is allowed to open a folder: everything else joins
    # the run this leaves open
    folder = dataio.create_run_dir(root, args.run_folder, fresh=True)
    emit({
        "run_folder": folder.name,
        "path": str(folder),
        "guidance": ("Pass --run-folder " + folder.name +
                     " on every profile and render command for this request. "
                     "Forget it and the command still writes into this same "
                     f"folder, so long as it runs within {dataio.OPEN_RUN_HOURS:g} "
                     "hours."),
    })


# --------------------------------------------------------------------------
def command_list(args: argparse.Namespace) -> None:
    root = Path(args.project_root).resolve()
    deps = dataio.load(require_geo=False)
    workbooks = []
    for path in dataio.find_excel_files(root):
        try:
            sheets = dataio.read_sheets(deps, path)
        except Exception as exc:  # pragma: no cover - depends on the workbook
            sheets = [f"<unreadable: {exc}>"]
        workbooks.append({"files": str(path.relative_to(root)), "sheet": sheets})

    boundary_root = dataio.shapefile_root(root)
    moved = dataio.migrate_legacy_layout(boundary_root)
    available = {}
    for name in dataio.countries(boundary_root):
        entry = {"tiers": [{k: v for k, v in tier.items() if not k.startswith("__")}
                          for tier in dataio.tiers(boundary_root, name)]}
        # Reading a country means opening its boundary files, which needs
        # geopandas. ``list`` is the command someone runs when something is
        # wrong, so it answers what it can without it rather than refusing.
        if deps.gpd is None:
            entry["profile"] = messages.text("list.geopandas-missing")
        else:
            try:
                reading = dataio.read_country(deps, boundary_root, name)
                entry.update({k: v for k, v in reading.items()
                              if k not in ("__from", "tiers")})
            except Exception as exc:        # a country that cannot be read is
                entry["unreadable"] = str(exc)   # reported, not fatal
        available[name] = entry

    # The per-tier paths are only meaningful when there is one country to be
    # meaningful about. With several installed, "country" above already says
    # what is there, and repeating the same refusal twice under a heading that
    # promises a file is noise dressed as an error.
    shapefiles = {}
    if len(available) == 1:
        for level in (dataio.COARSE, dataio.FINE):
            try:
                found = dataio.find_boundaries(root, level)
                # the boundaries may sit outside the project now, so a path
                # relative to it is not always expressible
                try:
                    shapefiles[level] = str(found.relative_to(root))
                except ValueError:
                    shapefiles[level] = str(found)
            except SystemExit as exc:
                shapefiles[level] = str(exc)

    emit({
        "project_root": str(root),
        "workbook": workbooks,
        "country": available,
        **({"layout_migrated": moved} if moved else {}),
        "shapefile": shapefiles,
        "missing_fonts": fonts.missing_files(),
        "remembered_choices": prefs._load(root / prefs.FOLDER / prefs.CHOICES),
    })


# --------------------------------------------------------------------------
def command_import(args: argparse.Namespace) -> None:
    """Take a file the user attached to the chat and put it where the rest of
    the skill can find it — then say what is inside it, in one step.

    Two things happen together on purpose. Copying alone would leave the agent
    to run ``survey`` next and the user waiting through two round trips for one
    obvious question; and a file that copies successfully but turns out to hold
    no place-name column is not really an import that succeeded.
    """
    root = Path(args.project_root).resolve()
    result = dataio.adopt_file(root, Path(args.file))

    # record it against the request when one is open, but never open one:
    # importing a file is not itself a reason to create an output folder
    run_name = args.run_folder or dataio.open_run(root)
    if run_name:
        _append_manifest(dataio.create_run_dir(root, run_name), "imports", result)

    survey = argparse.Namespace(project_root=args.project_root, excel=result["files"])
    emit({**result, "survey": _survey_payload(survey)})


def command_survey(args: argparse.Namespace) -> None:
    emit(_survey_payload(args))


def _survey_payload(args: argparse.Namespace) -> dict[str, Any]:
    """Which sheets in this workbook can become a map — before choosing one.

    Asking the user "which sheet?" before anyone knows which sheets are usable
    pushes the job back onto them, and they will pick the one whose name reads
    best. On a real export that is the pivot summary, not the data.

    Reads only a sample of each sheet: a place column is recognisable from a few
    hundred rows, and a workbook may hold sheets far too large to read in full
    just to decide whether they are worth reading.
    """
    from openpyxl import load_workbook

    root = Path(args.project_root).resolve()
    deps = dataio.load(require_geo=True)
    if args.excel:
        one = dataio.project_path(root, args.excel)
        if one is None or not one.exists():
            raise SystemExit(messages.text("error.workbook-not-found", file=args.excel))
        books = [one]
    else:
        # No file named: survey everything, so the question that follows can name
        # what each file holds. Asking "which file?" against a list of file names
        # is the same mistake as asking "which sheet?" — one level further out.
        books = dataio.find_excel_files(root)
        if not books:
            raise SystemExit(messages.text("survey.no-workbook"))

    country = getattr(args, "country", None)
    province_shapes = dataio.load_shapes(deps, root, dataio.COARSE, country=country)
    p_field = dataio.shape_fields(province_shapes, dataio.COARSE)["province"]
    affixes = dataio.name_affixes(
        dataio.read_country(deps, dataio.shapefile_root(root),
                            dataio.resolve_country(dataio.shapefile_root(root), country)))
    province_keys = {matching.normalize(v, affixes) for v in province_shapes[p_field]}
    province_keys |= {matching.normalize(n, affixes) for n in
                      crosswalk.build(province_shapes, name_field=p_field)}

    commune_shapes, c_fields = _fine_tier(deps, root, country)
    commune_keys = set()
    if commune_shapes is not None:
        commune_keys = {matching.normalize(v, affixes)
                        for v in commune_shapes[c_fields["commune"]]}

    limit = tabular.SURVEY_ROWS + tabular.MAX_HEADER_SCAN
    files = []
    for path in books:
        if dataio.is_text_table(path):
            try:
                sample, total = _sample_text_table(path, limit)
            except Exception as exc:
                files.append({"files": _short(path, root), "error": str(exc), "sheet": [],
                              "preferred": []})
                continue
            sheets = [_survey_sheet(dataio.SINGLE_SHEET, total, sample,
                                    province_keys, commune_keys, affixes)]
            files.append(_survey_file(path, root, sheets))
            continue
        try:
            book = load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:      # a workbook the reader cannot open at all
            files.append({"files": _short(path, root), "error": str(exc), "sheet": [],
                          "preferred": []})
            continue
        sheets = []
        try:
            for ws in book.worksheets:
                sample = []
                for row in ws.iter_rows(values_only=True):
                    sample.append(list(row))
                    if len(sample) >= limit:
                        break
                sheets.append(_survey_sheet(ws.title, int(ws.max_row or 0), sample,
                                            province_keys, commune_keys, affixes))
        finally:
            book.close()
        files.append(_survey_file(path, root, sheets))

    with_maps = [f for f in files if f["preferred"]]
    payload: dict[str, Any] = {
        "workbook": files,
        "preferred": [{"files": f["files"], "sheet": f["preferred"]} for f in with_maps],
        # A ready-made menu. Without it the agent lists what it finds in prose
        # and the reader has to copy a file name back — which is what happened
        # on the first real run, and it is a poor way to choose between nine
        # files. One number is the whole answer.
        "quick_pick": _quick_pick(files),
        "summary": (messages.text("survey.has-a-mappable-sheet", total=len(files),
                                  files=len(with_maps))
                    if len(files) > 1 else
                    (files[0].get("summary")
                     or messages.text("survey.unreadable"))),
        "note": messages.text("read.sampled-only", rows=tabular.SURVEY_ROWS),
    }
    if len(files) == 1:
        payload["sheet"] = files[0]["sheet"]      # unchanged shape for one file
    return payload


def _quick_pick(files: list[dict[str, Any]]) -> dict[str, Any]:
    """Every mappable sheet in the folder as a finished question.

    Sheets that cannot become a map are left out: offering them costs the reader
    a decision that only ends in a refusal. The last option is always to send a
    file instead — the folder holds fixtures and other people's work, and a menu
    with no way off it invites someone to map a table that is not theirs.
    """
    picks: list[dict[str, Any]] = []
    for f in files:
        for sheet in f.get("sheet", []):
            if not sheet.get("usable"):
                continue
            rows = sheet.get("estimated_rows") or 0
            level = sheet.get("suggested_level")
            picks.append({
                "number": len(picks) + 1,
                "label": f["files"],
                "description": messages.text(
                    "choice.file.description", sheet=sheet["sheet"],
                    rows=wording.count("table.row-count", "rows", rows),
                    level=messages.text({"commune": "tier.commune", "province": "tier.province"}
                                        .get(level, "table.not-applicable"))),
                "files": f["files"], "sheet": sheet["sheet"],
                "row_count": sheet.get("estimated_rows"), "level": level,
            })
    picks.append({"number": len(picks) + 1,
                  "label": messages.text("choice.file.upload.label"),
                  "description": messages.text("choice.file.upload.description"),
                  "files": None, "sheet": None})
    return {"question": messages.text("choice.file.question"), "choices": picks}


def _short(path: Path, root: Path) -> str:
    return str(path.relative_to(root) if path.is_relative_to(root) else path)


def _append_manifest(run_dir: Path, key: str, entry: dict[str, Any]) -> None:
    """Add one entry to the request's manifest, leaving the other keys alone.

    The manifest is the record of what a request did, and a request now does
    more than render: it may also have taken in a file the user attached. A
    corrupt or hand-edited manifest is rebuilt rather than allowed to stop a run
    that has already produced its maps.
    """
    path = run_dir / "run_manifest.json"
    manifest: dict[str, Any] = {}
    if path.exists():
        try:
            existing = json.loads(path.read_text(encoding="utf-8"))
            if isinstance(existing, dict):
                manifest = existing
        except (json.JSONDecodeError, OSError):
            pass
    manifest["folder"] = str(run_dir)
    if not isinstance(manifest.get(key), list):
        manifest[key] = []
    manifest[key].append(entry)
    dataio.write_json(path, manifest)


def _sample_text_table(path: Path, limit: int) -> tuple[list[list[Any]], int]:
    """The first rows of a delimited file, plus how many rows it really has.

    Read with the same dialect detection the full reader uses, so the survey
    cannot disagree with what ``profile`` will see a moment later.
    """
    dialect = dataio._text_dialect(path)
    sample: list[list[Any]] = []
    total = 0
    with open(path, encoding=dialect["encoding"], newline="") as handle:
        for row in csv.reader(handle, delimiter=dialect["sep"]):
            total += 1
            if len(sample) < limit:
                sample.append(list(row))
    return sample, total


def _survey_file(path: Path, root: Path, sheets: list[dict[str, Any]]) -> dict[str, Any]:
    usable = [s["sheet"] for s in sheets if s["usable"]]
    return {
        "files": _short(path, root), "sheet": sheets, "preferred": usable,
        "summary": (messages.text("survey.mappable-sheet", sheets=len(sheets),
                                  usable=len(usable), names=", ".join(usable))
                    if usable else
                    messages.text("survey.no-sheet-at-all", sheets=len(sheets))),
    }


def _several_tables(sample: list[list[Any]]) -> dict[str, Any]:
    """Whether this sheet holds more than one table, and where they start.

    Reported, never acted on. Read plainly, a sheet with a summary table above a
    budget table comes back as one table whose lower half has different columns
    — and no error. Which table the reader wants is their decision.
    """
    blocks = tabular.table_blocks(sample)
    if len(blocks) < 2:
        return {}
    return {
        "tables_in_sheet": len(blocks),
        "table_positions": [{"first_row": a + 1, "last_row": b + 1} for a, b in blocks],
        "caution": messages.text("read.several-tables", count=len(blocks)),
    }


def _tables_in_sheet(excel: Path, sheet: str | None) -> dict[str, Any]:
    """``_several_tables`` for one chosen sheet, for the profile to carry.

    ``survey`` reports a sheet holding two tables; ``profile`` did not, and the
    warning was lost the moment somebody named the sheet they wanted. Measured
    on a fixture with a second table appended below the first: the profile read
    straight through, and the second table's heading row — the literal string
    "Tỉnh/thành phố" — was fuzzy-matched onto Thanh Hóa at 88.9%.

    A delimited file cannot hold two tables the way a worksheet can. The suffix
    check is a short cut, not a correctness condition — ``load_workbook`` on a
    CSV raises and the answer would be the same either way.
    """
    if excel.suffix.lower() not in {".xlsx", ".xlsm", ".xls"}:
        return {}
    try:
        from openpyxl import load_workbook

        book = load_workbook(excel, read_only=True, data_only=True)
    except Exception:            # unreadable here is reported by the reader itself
        return {}
    try:
        ws = book[sheet] if sheet else book.worksheets[0]
        sample = []
        for row in ws.iter_rows(values_only=True):
            sample.append(list(row))
            if len(sample) >= tabular.SURVEY_ROWS:
                break
    except Exception:
        return {}
    finally:
        book.close()
    return _several_tables(sample)


def _survey_sheet(title: str, max_row: int, sample: list[list[Any]],
                  province_keys: set[str], commune_keys: set[str],
                  affixes) -> dict[str, Any]:
    """One sheet's verdict. Takes the title and row count rather than a
    worksheet object, so a delimited text file — which has neither sheets nor
    an openpyxl reader — goes through exactly the same judgement."""
    if not any(any(not tabular.is_blank(c) for c in row) for row in sample):
        return {"sheet": title, "usable": False, "estimated_rows": 0,
                "reason": messages.text("sheet-empty.reason"),
                "fix": messages.text("sheet-empty.fix")}

    start = tabular.header_row(sample)
    columns = [c for c in sample[start]] if start < len(sample) else []
    body = [r for r in sample[start + 1:] if any(not tabular.is_blank(c) for c in r)]
    names = [str(c).strip() if not tabular.is_blank(c) else f"Unnamed: {i}"
             for i, c in enumerate(columns)]

    places = tabular.place_columns(names, body, province_keys, commune_keys,
                                   affixes)
    place_column = places["commune"] or places["province"]
    total = max(int(max_row or 0) - start - 1, len(body))
    blocked = tabular.usability(names, len(body), place_column)

    out: dict[str, Any] = {
        "sheet": title,
        "usable": blocked is None,
        "estimated_rows": total,
        "column_count": len(names),
        "header_row": start + 1,
        "place_column": {"province": places["province"], "commune": places["commune"]},
        **_several_tables(sample),
        "suggested_level": ("commune" if places["commune"] else
                      "province" if places["province"] else None),
    }
    if blocked:
        out.update(blocked)
    return out


# --------------------------------------------------------------------------
def command_profile(args: argparse.Namespace) -> None:
    root = Path(args.project_root).resolve()
    deps = dataio.load(require_geo=True)
    excel = dataio.project_path(root, args.excel)
    if excel is None or not excel.exists():
        raise SystemExit(messages.text("error.workbook-not-found", file=args.excel))

    sheets = dataio.read_sheets(deps, excel)
    reading: list[dict[str, Any]] = []
    df = dataio.read_table(deps, excel, args.sheet, notes=reading)
    dictionary = dataio.read_data_dictionary(deps, excel, sheets)

    boundary_notes: list[dict[str, Any]] = []
    country = getattr(args, "country", None)
    province_shapes = dataio.load_shapes(deps, root, dataio.COARSE,
                                         notes=boundary_notes, country=country)
    p_fields = dataio.shape_fields(province_shapes, dataio.COARSE)
    province_names = [str(v) for v in province_shapes[p_fields["province"]]]

    commune_shapes, c_fields = _fine_tier(deps, root, country, boundary_notes)
    commune_names = ([str(v) for v in commune_shapes[c_fields["commune"]]]
                     if commune_shapes is not None else [])

    # Read once, from the profile, and handed to every place that matches a
    # name. Two lists that disagree miss every row, so there is one list.
    country_profile = dataio.read_country(
        deps, dataio.shapefile_root(root), country)
    affixes = dataio.name_affixes(country_profile)
    p_alias = dataio.alias_column(country_profile, dataio.COARSE)
    c_alias = dataio.alias_column(country_profile, dataio.FINE)

    admin_level = args.admin_level
    if admin_level in (None, "auto"):
        admin_level = _detect_admin_level(df, province_names, commune_names, affixes)

    report = profiling.build(deps, df, sheet=args.sheet, admin_level=admin_level,
                             province_names=province_names, commune_names=commune_names,
                             dictionary=dictionary, affixes=affixes)
    report["available_sheets"] = sheets
    report["remembered_choices"] = prefs.recall_choices(root, excel, args.sheet)
    tables = _tables_in_sheet(excel, args.sheet)
    if tables:
        report["many_tables_in_sheet"] = tables
        reading.append({"action": "many_tables", **tables})

    province_column = args.province_column or _first(report["place_column"]["province"])
    commune_column = args.commune_column or _first(report["place_column"]["commune"])
    report["suggested_province_column"] = province_column
    report["suggested_commune_column"] = commune_column

    place_column = commune_column if admin_level == "commune" else province_column
    report["sheet_reading"] = reading
    # A repaired codepage changes place names, so it is never left unsaid: the
    # user has to be able to tell a name the skill corrected from one they typed.
    if boundary_notes:
        report["codepage_repair"] = boundary_notes

    # Before anything else is said about this sheet: can it become a map at all?
    # Without this, a pivot table read with the wrong header row still came back
    # with a map option beside twenty columns called "Unnamed: 3".
    blocked = tabular.usability(list(df.columns), len(df), place_column)
    if blocked:
        report["usable"] = False
        report["map_options"] = []
        report["quality_warnings"].insert(0, guardrails._issue(
            "sheet-not-mappable", guardrails.CRITICAL,
            fmt={"why": blocked["reason"], "fix": blocked["fix"]},
            extra={k: v for k, v in blocked.items() if k not in ("reason", "fix")}))
        report["run_folder"] = dataio.create_run_dir(root, args.run_folder).name
        emit(report)
        return
    report["usable"] = True

    long_report = _long_format_report(df, report["column"], place_column)
    if long_report:
        report["long_form"] = long_report
        # the wide-format map options were computed on the assumption that a
        # column is a measure; here they are noise at best and misleading at worst
        report["map_options"] = []
        report["quality_warnings"].extend(long_report.pop("warnings", []))

    # the unit being mapped comes first: on commune data the province column is
    # the same value on every row and proves nothing
    places = ([commune_column, province_column] if admin_level == "commune"
              else [province_column, commune_column])
    report["sample"] = _sample_rows(df, report["column"], places, long_report)

    review: list[dict[str, Any]] = []
    if admin_level == "province" and province_column:
        index = _province_index(province_shapes, p_fields["province"], affixes,
                                p_alias)
        review = matching.review_province(
            [{"province": v} for v in df[province_column].tolist()], index)
    elif admin_level == "commune" and commune_column:
        index = _province_index(province_shapes, p_fields["province"], affixes,
                                p_alias)
        by_province = _commune_index_by_province(commune_shapes, c_fields, affixes,
                                                 c_alias)
        rows = [{"province": (df[province_column].iloc[i] if province_column else None),
                 "commune": df[commune_column].iloc[i]} for i in range(len(df))]
        review = matching.review_commune(rows, index, by_province,
                                         prefs.recall_overrides(root, "commune"))

    # every artefact of this request belongs to its run folder
    run_dir = dataio.create_run_dir(root, args.run_folder)
    summary = matching.summarize(review) if review else {}
    review_path = run_dir / review_filename(admin_level, excel, args.sheet)
    if review:
        dataio.write_csv(review_path, review)
    report["name_matching"] = {
        "summary": summary,
        "review_file": str(review_path) if review else None,
        "needs_review": matching.rows_needing_attention(review)[:25],
    }
    if summary:
        report["quality_warnings"].extend(guardrails.check_matching(summary))
        report["quality_warnings"].extend(
            guardrails.check_admin_level(summary, admin_level, commune_column))

    report["run_folder"] = run_dir.name
    report["run_folder"] = str(run_dir)
    profile_path = run_dir / "dataset_profile.json"
    dataio.write_json(profile_path, report)
    report["profile_file"] = str(profile_path)
    emit(report)


def _first(entries: list[dict[str, Any]]) -> str | None:
    return entries[0]["column"] if entries else None


LONG_SAMPLE = 5000
TOP_INDICATORS = 12

#: rows of one indicator read before deciding what kind of quantity it is —
#: enough to see whether the values are whole numbers or shares, cheap on a
#: sheet where one indicator can hold ten thousand rows
SEMANTIC_SAMPLE = 500

#: rows shown as evidence for what the agent claims the sheet is
SAMPLE_ROWS = 5
#: how many columns fit in a chat window before the table stops being readable
SAMPLE_COLUMNS = 7


def _sample_rows(df, columns: list[dict[str, Any]], place_columns: list[str],
                 long_report: dict[str, Any] | None) -> dict[str, Any]:
    """A few real rows, narrow enough to read, to back up the summary.

    A claim about a dataset is worth more with the dataset under it, but a
    twenty-one column table pasted into a chat window is unreadable and the
    reader skips it. So the columns that carry the meaning are chosen and the
    rest are counted, not hidden: the ellipsis column and the ellipsis row say
    plainly that there is more.
    """
    by_name = {c["column"]: c for c in columns}

    def take(names, into):
        for name in names:
            if name and name in df.columns and name not in into:
                into.append(name)

    # One place column is enough to show the grain; a second eats a slot that
    # the pinning columns need more. Those are what explain why one place
    # appears on many rows, so they come before the period columns.
    wanted: list[str] = []
    take([c for c in place_columns if c][:1], wanted)

    if long_report:
        take([long_report.get("indicator_column"), long_report.get("value_column")], wanted)
        take([a["column"] for a in long_report.get("columns_to_pin", [])][:2], wanted)
        take(long_report.get("period_column", [])[:2], wanted)
    else:
        for info in columns:
            if len(wanted) >= SAMPLE_COLUMNS:
                break
            if info["column"] in wanted or not info.get("mappable"):
                continue
            if info["semantic"] in longform.MEASURE_SEMANTICS:
                wanted.append(info["column"])

    for info in columns:                       # top up with whatever is left
        if len(wanted) >= SAMPLE_COLUMNS:
            break
        if info["column"] not in wanted:
            wanted.append(info["column"])
    wanted = wanted[:SAMPLE_COLUMNS]

    hidden = [str(c) for c in df.columns if c not in wanted]
    rows = []
    for _, row in df.head(SAMPLE_ROWS).iterrows():
        rows.append([sem.format_value(row[c], by_name.get(c, {}))
                     if by_name.get(c, {}).get("semantic") in longform.MEASURE_SEMANTICS
                     else ("" if tabular.is_blank(row[c]) else str(row[c]))
                     for c in wanted])
    return {
        "column": wanted,
        "rows": rows,
        "hidden_columns": len(hidden),
        "hidden_column_names": hidden,
        "remaining_rows": max(len(df) - len(rows), 0),
        "display": ("Add a '…' column at the end if columns are hidden, and a "
                    "'…' row below if the table runs longer, each with the "
                    "count beside it."),
    }


def _apply_where(df, expressions):
    """Keep only the rows the user asked for, comparing as text.

    A filter that matches nothing stops the run and lists the values that do
    exist. Silently drawing an empty map because a value was typed 'Q02' is the
    kind of failure that costs an hour before anyone thinks to check the spelling.
    """
    pairs = longform.parse_where(expressions or [])
    if not pairs:
        return df, None
    for column, wanted in pairs:
        if column not in df.columns:
            raise SystemExit(messages.text(
                "error.where-unknown-column", column=repr(column),
                available=", ".join(map(str, df.columns))))
        near = longform.unknown_values(column, wanted, df[column].tolist())
        if near is not None:
            raise SystemExit(messages.text(
                "error.where-no-rows", column=column, value=wanted,
                near=", ".join(repr(v) for v in near)))
        df = df[df[column].astype(str).str.strip() == wanted]
    df = df.reset_index(drop=True)
    if df.empty:
        raise SystemExit(messages.text("error.where-empty-result"))
    return df, longform.describe_filters(pairs)


def _units_shown(frame, has_fill: bool, has_symbol: bool) -> int:
    """Units the finished map shows something for, on any channel.

    A unit counts once whether it is coloured, circled or both — the question
    being answered is "how much of this map carries data", not "how many values
    are there".
    """
    mask = None
    for column, present in (("__value", has_fill), ("__symbol", has_symbol)):
        if not present or column not in frame.columns:
            continue
        here = frame[column].notna()
        mask = here if mask is None else (mask | here)
    return 0 if mask is None else int(mask.sum())


#: separates numerator from denominator in ``--layer "TX_PVLS Num / TX_PVLS Den"``
RATIO_SEPARATOR = "/"

#: On a wide sheet the separator has to carry spaces. Indicator codes hold no
#: slash, so a long sheet can split on a bare one; column headings hold plenty —
#: ``Tỷ suất ca mới/100.000 dân`` and ``Status/Result`` are both real — and
#: splitting those would cut a column in half and then report it missing.
WIDE_RATIO_SEPARATOR = " / "

#: separates an indicator from its own pins, as in
#: ``--layer "HTS_TST_POS|Status/Result=Positive"``. A bar rather than a slash,
#: because column names in real exports contain slashes ("Status/Result") and
#: indicator codes do not contain bars.
PIN_SEPARATOR = "|"


def split_wide_ratio(text: Any) -> tuple[str, str | None]:
    """``"A / B"`` -> ``("A", "B")``; anything else -> ``(text, None)``.

    Spaces around the slash are required, and that is the whole reason this is
    a function rather than a ``split``: a wide sheet's headings contain slashes
    of their own.
    """
    raw = str(text or "")
    if WIDE_RATIO_SEPARATOR not in raw:
        return raw, None
    left, _, right = raw.partition(WIDE_RATIO_SEPARATOR)
    left, right = left.strip(), right.strip()
    return (left, right) if left and right else (raw, None)


def build_wide_ratio(joined, by_name, spec: Any) -> str | None:
    """Add ``A ÷ B (%)`` to the frame, or return None if ``spec`` is not one.

    Per unit the two columns are summed and divided after, never averaged from
    row-level ratios — the same rule the long sheet follows, and the same reason:
    a mean of ratios weights a commune of two hundred like a city of two million.
    """
    num, den = split_wide_ratio(spec)
    if den is None:
        return None
    if num not in joined.columns or den not in joined.columns:
        raise SystemExit(messages.text(
            "error.layer-unknown-column",
            column=repr(num if num not in joined.columns else den),
            available=", ".join(map(str, joined.columns))))
    name = f"{num} ÷ {den} (%)"
    bottom = joined[den].where(joined[den] != 0)
    joined[name] = joined[num] / bottom * 100.0
    by_name[name] = sem._pack(sem.PERCENT, name, "percent", scale="percent")
    return name


def _layer_requests(args, deps, frame) -> list[dict[str, Any]]:
    """What each ``--layer`` refers to, on a wide sheet and on a long one.

    Wide: a column name, and its semantic is already known from the heading.
    Long: a value inside ``--indicator-column``, because every indicator shares
    one numeric column — asking for a column name there is meaningless, and
    used to be a dead end with a good error message and no way forward.
    """
    axis = args.indicator_column
    if not axis:
        known = {c["column"]: c for c in profiling.describe_columns(deps, frame, None)}
        out = []
        for name in args.layer:
            num, den = split_wide_ratio(name)
            if den is not None:
                for side in (num, den):
                    if side not in known:
                        raise SystemExit(messages.text(
                            "error.layer-unknown-column", column=repr(side),
                            available=", ".join(map(str, frame.columns))))
                # a quotient is normalised by construction, so it belongs on the
                # fill channel whatever the two columns it came from were
                out.append({"name": f"{num} ÷ {den} (%)", "semantic": sem.PERCENT,
                            "column": name})
                continue
            if name not in known:
                raise SystemExit(messages.text(
                    "error.layer-unknown-column", column=repr(name),
                    available=", ".join(map(str, frame.columns))))
            out.append({"name": name, "semantic": known[name]["semantic"], "column": name})
        return out

    if axis not in frame.columns:
        raise SystemExit(messages.text(
            "error.indicator-column-unknown", column=repr(axis),
            available=", ".join(map(str, frame.columns))))
    if not args.value_column:
        raise SystemExit(messages.text("error.long-table-needs-value-column"))

    codes = frame[axis].astype(str).str.strip()
    present = set(codes)

    def require(value: str) -> None:
        if value in present:
            return
        near = longform.unknown_values(axis, value, frame[axis].tolist()) or []
        raise SystemExit(messages.text(
            "error.layer-value-not-in-column", value=value, column=repr(axis),
            near=", ".join(repr(v) for v in near)))

    out = []
    for raw in args.layer:
        # pins first: a column name may contain a slash ("Status/Result"), so
        # splitting on the ratio separator before this would misread the pin
        head, *pins = (part.strip() for part in raw.split(PIN_SEPARATOR))
        if RATIO_SEPARATOR in head:
            num, den = (p.strip() for p in head.split(RATIO_SEPARATOR, 1))
            require(num)
            require(den)
            # a quotient is normalised by construction, so it belongs to the fill
            out.append({"name": f"{num} ÷ {den} (%)", "semantic": sem.PERCENT,
                        "numerator": num, "denominator": den, "slice": pins})
            continue
        require(head)
        rows = frame[codes == head]
        for column, value in longform.parse_where(pins):
            if column not in rows.columns:
                raise SystemExit(messages.text("error.layer-slice-unknown-column",
                                               layer=head, column=repr(column)))
            rows = rows[rows[column].astype(str).str.strip() == value]
        info = indicator_semantic(rows[args.value_column].tolist(), head)
        out.append({"name": head, "semantic": info["semantic"], "indicator": head, "slice": pins})
    return out


def _render_layers(args: argparse.Namespace) -> None:
    """Draw however many maps the requested variables need.

    A map holds two quantitative channels. Asking for three variables is not an
    error and should not be answered with one — the surplus goes to a second
    map in the same folder, behind the same picker on the interactive page.

    Each map is rendered by a full pass of ``command_render`` rather than by
    looping inside it. That costs a re-read of the workbook, and buys the one
    thing that matters here: every map computes its own class breaks and its own
    circle scale, which is correct because two different variables have no
    business sharing either.
    """
    root = Path(args.project_root).resolve()
    deps = dataio.load(require_geo=False)
    excel = dataio.project_path(root, args.excel)
    frame, _ = _apply_where(dataio.read_table(deps, excel, args.sheet), args.where)

    requests = _layer_requests(args, deps, frame)
    plan = layers.allocate(requests)
    if not plan["maps"]:
        raise SystemExit(messages.text("error.no-mappable-variable")
                         + " ".join(r["why"] for r in plan["unplaced"]))

    # one folder for the whole set, resolved once so every pass writes together
    run_folder = args.run_folder or dataio.create_run_dir(root).name
    emit({"layer_plan": layers.summary_lines(plan),
          "why_split": plan.get("why_split"),
          "unplaced": plan["unplaced"],
          "layer_warnings": layers.conflicts(requests),
          "folder": run_folder,
          "maps_to_draw": len(plan["maps"])})

    for item in plan["maps"]:
        fill, symbol = item["fill"], item["symbol"]
        step = argparse.Namespace(**vars(args))
        step.layer = None
        step.run_folder = run_folder
        step.map_type = item["kind"]
        if args.indicator_column:
            # each channel names its own slice; the numeric column stays shared
            step.numerator = (fill or {}).get("numerator")
            step.denominator = (fill or {}).get("denominator")
            step.fill_indicator = (fill or {}).get("indicator")
            step.symbol_indicator = (symbol or {}).get("indicator")
            step.fill_where = (fill or {}).get("slice") or args.fill_where
            step.symbol_where = (symbol or {}).get("slice") or args.symbol_where
            step.symbol_column = None
        else:
            step.value_column = fill["column"] if fill else None
            step.symbol_column = symbol["column"] if symbol else None
        # the file name is built from the title, so a fixed title across several
        # maps would have each pass overwrite the one before it
        if len(plan["maps"]) > 1:
            lead = (fill or symbol)["name"]
            step.title = f"{args.title} — {lead}" if args.title else lead
        command_render(step)


def indicator_semantic(values: list[Any], name: str) -> dict[str, Any]:
    """What kind of quantity one indicator's own rows hold.

    On a long sheet the column heading says nothing — every indicator shares the
    same ``Value`` column — so the semantic has to be read off the numbers.
    """
    clean = [v for v in values if v is not None and v == v][:SEMANTIC_SAMPLE]
    return sem.infer(name, clean, True)


def _drawn_table(run_dir: Path, base: str, frame, name_field: str,
                 value_column: str | None, value_info: dict[str, Any],
                 symbol_info: dict[str, Any], args, bins) -> str | None:
    """The numbers the map was drawn from, as a CSV beside the image.

    Not the input sheet: this is what survived name matching, filtering and
    aggregation — the values actually behind the colours and the circles.
    Anybody checking a figure in a report needs that table, and reconstructing
    it from the workbook means repeating every step by hand and hoping.

    Both the raw value and the formatted one are written: the first to compute
    with, the second to check against what the reader sees on the plate.
    """
    if name_field not in frame.columns:
        return None
    rows: list[dict[str, Any]] = []
    lang = args.language
    # the same number of decimals the plate's labels use. Formatting the table
    # independently made the map say 99,74% and the table 99,7% about one value
    # — the column exists to be checked against the plate, so it has to agree
    # with it down to the last digit.
    decimals = classify.label_decimals(bins)
    for _, row in frame.iterrows():
        # Headings follow --language, so the Vietnamese edition of a map gets a
        # Vietnamese table and the English edition an English one. The user's own
        # column name wins wherever there is one: that heading is their data.
        entry: dict[str, Any] = {i18n.t(lang, "csv_unit"): row[name_field]}
        if "__value" in frame.columns:
            raw = row["__value"]
            entry[value_column or i18n.t(lang, "csv_value")] = "" if raw != raw else raw
            entry[i18n.t(lang, "csv_formatted")] = (
                "" if raw != raw else sem.format_value(raw, value_info,
                                                       decimals=decimals, lang=lang))
        if "__symbol" in frame.columns:
            raw = row["__symbol"]
            entry[args.symbol_column or i18n.t(lang, "csv_symbol")] = (
                "" if raw != raw else raw)
            entry[i18n.t(lang, "csv_symbol_formatted")] = (
                "" if raw != raw else sem.format_value(raw, symbol_info, decimals=0,
                                                       lang=lang))
        rows.append(entry)
    if not rows:
        return None
    path = run_dir / f"{base}_data.csv"
    dataio.write_csv(path, rows)
    return str(path)


def _dress_points(points: dict[str, Any], rows, args, by_name) -> None:
    """Give the dots a colour and a size, when the user named columns for them.

    Same rule as the area map, for the same reason: a category chooses colour, a
    magnitude chooses size. Sizing dots by a category would invent an order the
    data does not have, and colouring them by a count would ask a reader to
    judge quantity from hue.
    """
    colour_col = args.point_color_column
    if colour_col:
        if colour_col not in rows.columns:
            raise SystemExit(messages.text("error.point-colour-column-unknown",
                                           column=repr(colour_col)))
        labels = [str(v) for v in rows[colour_col]]
        cats, mapping = classify.category_colours(labels)
        points["colours"] = [mapping[v] for v in labels]
        points["legend_pairs"] = [(mapping[c], c) for c in cats]
        points["fill_by"] = colour_col

    size_col = args.point_size_column
    if size_col:
        if size_col not in rows.columns:
            raise SystemExit(messages.text("error.point-size-column-unknown",
                                           column=repr(size_col)))
        values = [None if v != v else float(v) for v in rows[size_col]]
        finite = [v for v in values if v is not None and v > 0]
        vmax = max(finite) if finite else 1.0
        s_max, s_min = furniture.SYMBOL_MAX_PT ** 2, furniture.SYMBOL_MIN_PT ** 2
        # the floor is a display minimum and stays out of the scale, or the key
        # would stop being true for the smallest dots
        points["sizes"] = [s_min if not v else max(s_min, s_max * (v / vmax))
                           for v in values]
        points["size_scale"] = {"max_value": vmax,
                                "title": args.symbol_legend_title or size_col}
        points["size_info"] = by_name.get(size_col, {})
        points["size_by"] = size_col


def _long_keys(args) -> list[str]:
    """What one row of the reduced frame stands for.

    A still map wants one row per unit. A map over time wants one row per unit
    **per period**, or every frame of the film shows the same picture: the sum
    of all of them. Reducing to the unit alone is what made a column of four
    quarters arrive at the animation as one.
    """
    keys = ["__shape_id"]
    if args.animate and args.period_column:
        keys.append(args.period_column)
    return keys


def _spread(frame, totals, keys):
    """Put a grouped total back on the frame, on the same keys it was grouped by."""
    import pandas as pd

    if len(keys) == 1:
        return frame[keys[0]].map(totals)
    return pd.Series(pd.MultiIndex.from_frame(frame[keys]).map(totals),
                     index=frame.index)


def _build_long_columns(args, joined, by_name):
    """Columns the sheet does not have: one per channel, each from its own rows.

    On a long sheet "viral suppression" and "patients on treatment" are not two
    columns, they are two sets of rows. So each channel of the map has to be
    built from its own slice, and neither can be read off a column that exists
    in the file.

    TX_PVLS Num and TX_PVLS Den are separate rows, and neither is a map. Their
    quotient is viral suppression, and the division happens **after** summing
    each side within a unit — dividing row by row and averaging the quotients is
    a different, wrong number whenever the units differ in size.
    """
    axis = args.indicator_column or args.ratio_column
    if not axis:
        raise SystemExit(messages.text("error.indicator-column-required"))
    if axis not in joined.columns:
        raise SystemExit(messages.text(
            "error.indicator-column-unknown", column=repr(axis),
            available=", ".join(map(str, joined.columns))))
    if not args.value_column:
        raise SystemExit(messages.text("error.long-table-needs-value-column"))

    codes = joined[axis].astype(str).str.strip()

    def rows_for(label: str, wanted: str, pins: list[str] | None = None):
        """The rows of one indicator, pinned on its own terms.

        Two indicators on one map can need different pins on the same column:
        in a PEPFAR extract TX_CURR carries a pre-computed 'Total' row beside
        its detail, while HTS_TST_POS has no Total at all and is pinned on
        'Positive'. A single --where would have to choose one and lose the
        other, so each channel pins its own slice here instead.
        """
        part = joined[codes == wanted]
        if part.empty:
            near = longform.unknown_values(axis, wanted, joined[axis].tolist()) or []
            raise SystemExit(messages.text(
                "error.no-rows-for-indicator", label=label, value=wanted,
                near=", ".join(repr(v) for v in near)))
        for column, value in longform.parse_where(pins or []):
            if column not in part.columns:
                raise SystemExit(messages.text("error.slice-unknown-column",
                                               indicator=wanted, column=repr(column)))
            pinned = part[part[column].astype(str).str.strip() == value]
            if pinned.empty:
                have = sorted({str(v).strip() for v in part[column].tolist()})[:8]
                raise SystemExit(messages.text(
                    "error.slice-no-rows", column=column, value=value,
                    indicator=wanted, near=", ".join(map(repr, have))))
            part = pinned
        return part

    keys = _long_keys(args)
    frame = joined.drop_duplicates(keys).copy()
    note: dict[str, Any] = {"indicator_column": axis}
    if len(keys) > 1:
        note["kept_per_period"] = args.period_column
    name = None

    if args.numerator or args.denominator:
        if not (args.numerator and args.denominator):
            raise SystemExit(messages.text("error.ratio-needs-both"))
        top = rows_for(messages.fragment("numerator"), args.numerator, args.fill_where) \
            .groupby(keys)[args.value_column].sum()
        bottom = rows_for(messages.fragment("denominator"), args.denominator, args.fill_where) \
            .groupby(keys)[args.value_column].sum()
        share = (top / bottom.replace(0, float("nan")) * 100.0).dropna()
        name = f"{args.numerator} ÷ {args.denominator} (%)"
        frame[name] = _spread(frame, share, keys)
        frame = frame[frame[name].notna()]
        by_name[name] = sem._pack(sem.PERCENT, name, "percent", scale="percent")
        note.update({"numerator": args.numerator, "denominator": args.denominator,
                     "fill_slice": list(args.fill_where or []),
                     "computable_units": int(frame["__shape_id"].nunique()),
                     "zero_denominator_units": len(set(bottom.index) - set(share.index)),
                     "method": "The numerator and the denominator are summed within each "
                               "unit and divided after, rather than averaging "
                               "the ratios."})
    elif args.fill_indicator:
        rows = rows_for("fill", args.fill_indicator, args.fill_where)
        totals = rows.groupby(keys)[args.value_column].sum()
        name = str(args.fill_indicator)
        frame[name] = _spread(frame, totals, keys)
        by_name[name] = indicator_semantic(rows[args.value_column].tolist(), name)
        note["fill"] = {"indicator": name,
                        "unit_count": int(frame.loc[frame[name].notna(),
                                                    "__shape_id"].nunique()),
                            "total": float(totals.sum()),
                            "slice": list(args.fill_where or [])}

    # The circles come from a different slice of the same sheet. Without this the
    # symbol column would be read off whatever single row survived the
    # de-duplication above — a real value from the wrong indicator, drawn at a
    # believable size, with nothing on the map to say so.
    if args.symbol_indicator:
        rows = rows_for("symbol", args.symbol_indicator, args.symbol_where)
        totals = rows.groupby(keys)[args.value_column].sum()
        symbol_name = str(args.symbol_indicator)
        frame[symbol_name] = _spread(frame, totals, keys)
        by_name[symbol_name] = sem._pack(sem.COUNT, symbol_name, "count", integer=True)
        args.symbol_column = symbol_name
        note["symbol"] = {"indicator": symbol_name,
                          "unit_count": int(frame.loc[frame[symbol_name].notna(),
                                                      "__shape_id"].nunique()),
                             "total": float(totals.sum()),
                             "slice": list(args.symbol_where or [])}
    return name, frame, note


def _long_format_report(df, columns: list[dict[str, Any]],
                        place_column: str | None) -> dict[str, Any] | None:
    """What an analyst would need to know before choosing anything.

    A long sheet cannot be described by listing its columns: the map-worthy
    question is which indicator, over which period, from which disaggregation.
    This works that out from the data and, just as importantly, names the
    columns that must be pinned before any number is trustworthy.
    """
    if not place_column or place_column not in df.columns:
        return None
    shape = longform.looks_long(columns, len(df), int(df[place_column].nunique()))
    if not shape:
        return None

    samples = {c: df[c].dropna().astype(str).head(LONG_SAMPLE).tolist()
               for c in df.columns}
    axis = longform.indicator_axis(columns, samples)
    value_column = shape["value_column"]
    time_columns = [c["column"] for c in columns if c.get("semantic") == sem.TIME]
    period_column = time_columns[0] if time_columns else None

    out: dict[str, Any] = {
        **shape,
        "indicator_column": axis,
        "period_column": time_columns,
        "columns_to_pin": [],
        "indicator": [],
        "ratio_pairs": [],
        "usage": None,
        "warnings": [],
    }

    axes = {c: df[c].tolist() for c in df.columns
            if c not in {place_column, value_column} and df[c].nunique() > 1}
    risky = longform.double_counting_axes(df[place_column].tolist(), axes)
    out["columns_to_pin"] = risky
    warning = longform.pin_warning(risky, int(df[place_column].nunique()))
    if warning:
        out["warnings"].append(guardrails._issue(
            "long-table-double-count", guardrails.CRITICAL, fmt={"why": warning}))

    if not axis:
        return out

    # plain lists, not Series: a filtered Series keeps the original index, so
    # positional access silently becomes a label lookup
    out["indicator"] = longform.indicator_slices(
        df[axis].tolist(), df[place_column].tolist(), df[value_column].tolist(),
        df[period_column].tolist() if period_column else None)[:TOP_INDICATORS]
    out["ratio_pairs"] = longform.ratio_pairs(df[axis].dropna().unique())

    # Naming the dangerous columns is not enough: the agent still has to know
    # which value to pin, and the data can answer that better than a guess.
    # Time first. A stock indicator like "patients currently on treatment" is
    # not additive across quarters: summing six reporting periods counted every
    # patient six times and produced 433.681 where the quarter alone is 49.706.
    pin_columns = time_columns + [a["column"] for a in risky if a["column"] not in time_columns]
    for entry in out["indicator"]:
        entry.update(_recommend_slice(df, axis, entry["indicator"], place_column,
                                      value_column, pin_columns, time_columns))

    if out["indicator"]:
        out["usage"] = out["indicator"][0]["command"]
    return out


def _recommend_slice(df, axis: str, indicator: str, place_column: str,
                     value_column: str, pin_columns: list[str],
                     time_columns: list[str]) -> dict[str, Any]:
    """For one indicator: which value to pin on each column, and why.

    Measured on the indicator's own rows, because the right disaggregation for
    TX_CURR is not the right one for HTS_TST — a recommendation taken from the
    whole sheet would be confidently wrong for most of the list.
    """
    rows = df[df[axis].astype(str).str.strip() == indicator]
    if rows.empty:
        return {"suggested_slices": [], "command": None, "unpinned": []}

    chosen: list[dict[str, Any]] = []
    kept = rows
    for column in pin_columns:
        if column not in rows.columns:
            continue
        if column in time_columns:
            pick, same = _latest_period(kept, column, place_column, value_column), []
        else:
            options = longform.pin_options(kept[column].tolist(),
                                           kept[place_column].tolist(),
                                           kept[value_column].tolist())
            pick = longform.recommend_pin(options)
            same = longform.duplicated_totals(options)
        if not pick:
            continue
        chosen.append({
            "column": column, "pinned": pick["value"], "why": pick["why"],
            "unit_count": pick["unit_count"], "total": pick["total"],
            "alternatives": pick["alternatives"],
            "same_total_values": same,
        })
        kept = kept[kept[column].astype(str).str.strip() == pick["value"]]

    # Whatever still splits a place after all that is a decision, not a default:
    # say so instead of summing across it and reporting the total as settled.
    leftover = {c: kept[c].tolist() for c in kept.columns
                if c not in {place_column, value_column, axis}
                and c not in {p["column"] for p in chosen} and kept[c].nunique() > 1}
    remaining = [a["column"] for a in
                 longform.varying_axes(kept[place_column].tolist(), leftover)]

    pins = " ".join(f'--where "{c["column"]}={c["pinned"]}"' for c in chosen)
    command = (f'--value-column "{value_column}" --where "{axis}={indicator}" '
               f'{pins}').strip()
    return {"suggested_slices": chosen,
            "total_after_pinning": chosen[-1]["total"] if chosen else None,
            "unpinned": remaining[:6],
            "command": command}


def _latest_period(rows, column: str, place_column: str,
                   value_column: str) -> dict[str, Any] | None:
    """The most recent value of a time column, in real chronological order.

    Alphabetical order puts Q4 before Q1 of the next year and 'Target' after
    both, so the ordering has to come from the period reader, not from sorting
    the strings.
    """
    values = [v for v in rows[column].dropna().tolist()]
    if not values:
        return None
    ordered = period_utils.ordered(values)
    if not ordered:
        return None
    latest = str(ordered[-1]).strip()
    hit = rows[rows[column].astype(str).str.strip() == latest]
    return {
        "value": latest,
        "unit_count": int(hit[place_column].nunique()),
        "total": round(float(deps_sum(hit[value_column])), 1),
        "why": messages.text("longform.latest-period", count=len(ordered)),
        "alternatives": [str(p) for p in reversed(ordered[:-1])][:6],
    }


def _points_per_unit(deps, frame, points: dict[str, Any] | None) -> dict[str, int]:
    """How many plotted locations fall inside each administrative unit.

    On a point map the mapped quantity belongs to the dots, not to the areas, so
    a unit's detail card would otherwise have nothing to report but its own size.
    This is a count of what is actually drawn on that map — derived from the same
    coordinates — rather than an estimate.
    """
    if not points or not points.get("x"):
        return {}
    located = deps.gpd.GeoDataFrame(
        geometry=deps.gpd.points_from_xy(points["x"], points["y"]),
        crs=frame.crs)
    hit = deps.gpd.sjoin(located, frame[["__shape_id", "geometry"]],
                         how="inner", predicate="within")
    return {str(int(sid)): int(n) for sid, n in hit["__shape_id"].value_counts().items()}


def deps_sum(series) -> float:
    total = series.sum()
    try:
        return float(total)
    except (TypeError, ValueError):
        return 0.0


def review_filename(admin_level: str, excel: Path, sheet: str | None) -> str:
    """One review per dataset *and* level.

    A single request may map two different workbooks — a province table and a
    commune table — into the same folder. One shared file name meant the second
    render picked up the first render's review and failed on a missing column.
    """
    slug = dataio.slugify(f"{Path(excel).stem} {sheet or ''}", fallback="dataset")
    return f"match_review_{admin_level}_{slug}.csv"


def _review_is_usable(review: list[dict[str, Any]], admin_level: str) -> bool:
    if not review:
        return False
    needed = {"dataset_province"}
    if admin_level == "commune":
        needed.add("dataset_commune")
    return needed.issubset(review[0].keys())


# --------------------------------------------------------------------------
def command_fix_match(args: argparse.Namespace) -> None:
    root = Path(args.project_root).resolve()
    deps = dataio.load(require_geo=True)
    gdf = dataio.load_shapes(deps, root, args.admin_level,
                             country=getattr(args, "country", None))
    fields = dataio.shape_fields(gdf, args.admin_level)
    field = fields["commune"] if args.admin_level == "commune" else fields["province"]
    row = gdf[gdf["__shape_id"] == int(args.shape_id)]
    if row.empty:
        raise SystemExit(messages.text("error.no-unit-with-shape-id",
                                       shape_id=args.shape_id))
    path = prefs.remember_override(root, args.admin_level, args.province, args.name,
                                   int(args.shape_id), str(row.iloc[0][field]))
    emit({"remembered": {"name_in_table": args.name,
                         "matched_to": str(row.iloc[0][field]),
                         "shape_id": int(args.shape_id)},
          "files": str(path)})


# --------------------------------------------------------------------------
def _contexts(args, shapes, fields, review, admin_level):
    """Decide what gets drawn: one national map, one province, or a series."""
    matched = [r for r in review if str(r.get("shape_id", "")) != ""]
    ids = {int(r["shape_id"]) for r in matched}
    scope = args.map_scope or "auto"

    if admin_level == "province":
        frame = shapes if scope != "matched-only" else shapes[shapes["__shape_id"].isin(ids)]
        return "national", [{"name": "national", "frame": frame, "locator": None}]

    provinces = sorted({r.get("matched_province", "") for r in matched
                        if r.get("matched_province")})
    if not provinces:
        # nothing was matched by name (e.g. a coordinate-only point map)
        return "national", [{"name": "national", "frame": shapes, "locator": None}]
    if scope == "matched-only":
        return "matched-only", [{"name": "units with data",
                                 "frame": shapes[shapes["__shape_id"].isin(ids)],
                                 "locator": provinces[0] if len(provinces) == 1 else None}]
    if scope == "national" or (scope == "auto"
                               and len(provinces) > args.province_series_threshold):
        return "national", [{"name": "national", "frame": shapes, "locator": None}]
    if scope == "single-province" or len(provinces) == 1:
        name = provinces[0] if provinces else ""
        return "single-province", [{"name": name,
                                    "frame": shapes[shapes[fields["province"]] == name],
                                    "locator": name}]
    return "province-series", [
        {"name": name, "frame": shapes[shapes[fields["province"]] == name], "locator": name}
        for name in provinces
    ]


def _command_line() -> str:
    """This invocation, as the agent would type it again."""
    return " ".join(shlex.quote(a) for a in sys.argv)


#: Settings the person is meant to decide, not the skill. When one of these was
#: not supplied on the command line its question goes into ``must_ask``, so the
#: agent asks about that rather than presenting a guess as a decision. The map's
#: language is here because a real run inferred it from the chat and never asked
#: — the reader may well want a Vietnamese map while writing in English.
THEIRS_TO_CHOOSE = ("language", "layout")


def _among(value_column, args, scope) -> dict[str, tuple[str, ...]]:
    """Alternatives that only apply to some tables, worked out from this one.

    :mod:`wording` will offer a menu for anything, and for most settings that is
    safe. These two are not: whether another map type or another framing is
    available depends on the columns and on where the rows fall, so the pool is
    narrowed here rather than guessed at by the agent.

    ``matched-only`` is never among them. It is a real flag and it stays
    reachable when a person asks for it by name, but a menu that offers it turns
    "we surveyed 12 of 34 provinces" into a picture of a country with 12
    provinces, and no reader of that picture would know.
    """
    among: dict[str, tuple[str, ...]] = {}
    if value_column and args.symbol_column:
        # both channels are filled, so all three area maps can be drawn from
        # exactly these columns
        among["map_type"] = ("choropleth-symbol", "choropleth", "graduated-symbol")
    if scope in ("national", "province-series"):
        among["map_scope"] = ("national", "province-series")
    return among


def _title_ingredients(args, joined, value_column, scope, prepared,
                       country_name) -> dict[str, Any]:
    """What the engine knows that a title ought to mention.

    Not a title — the engine has no standing to name somebody else's figures.
    These are the facts a good one is built from, and the plate that prompted
    this was missing every one of them: it drew two columns and named one, over
    a country and a year it never said.
    """
    drawn = [c for c in (value_column, args.symbol_column) if c]
    if scope == "national":
        place = country_name or ""
    else:
        place = ", ".join(str(c["name"]) for c in prepared)
    periods = []
    if args.period_column and args.period_column in joined.columns:
        periods = [str(p) for p in period_utils.ordered(joined[args.period_column])]
    return {"columns": drawn, "place": place, "periods": periods}


def _plan(args, excel, joined, value_column, scope, prepared, method, bins,
          country_name=None):
    """The numbered table the person agrees to, and the settings it stands for.

    Every value is written in the language of the conversation and in ordinary
    words — a table reading ``choropleth-symbol`` and ``weighted-mean`` can only
    be agreed to, never weighed. Rows the reader can still change carry their
    own question and their own alternatives, ready to hand to a picker.

    The table and the hash are built from the same values on purpose: what the
    reader saw is exactly what unlocks the drawing.
    """
    maps = ", ".join(str(c["name"]) for c in prepared)
    auto = messages.text("table.chosen-by-the-skill")
    chosen = getattr(args, "chosen_explicitly", set())
    among = _among(value_column, args, scope)
    unknown = messages.text("table.not-applicable")

    # a CSV has no sheet, and "sheet None" is worse jargon than any flag value
    sheet = f" › sheet {args.sheet}" if args.sheet else ""
    counted = wording.count("table.plate-count", "maps", len(prepared))
    # "One national map — national" says the same thing twice; the place names
    # only earn their space when they are not already the scope's own name
    reach = wording.label("map_scope", scope)
    if len(prepared) > 1:
        reach += f" — {counted}: {maps}"
    elif scope != "national":
        reach += f" — {maps}"

    # (row name, what it says, which setting it stands for)
    rows: list[tuple[str, str, str | None]] = [
        ("data", f"{Path(excel).name}{sheet} "
                    f"({wording.count('table.row-count', 'rows', len(joined))})", None),
    ]
    if args.where:
        rows.append(("data-slice", " · ".join(args.where), None))
    rows += [
        ("map-kind", wording.label("map_type", args.map_type), "map_type"),
        ("coloured-by", value_column or "—", None),
    ]
    if args.symbol_column:
        rows.append(("circles-by", args.symbol_column, None))

    # The wording on the plate, shown as its own rows so the person can weigh
    # it. A value that fell back to a column name says so: "Tỷ lệ dương tính
    # (%)" is a heading in a spreadsheet, not a title on a map.
    from_column = messages.text("table.from-the-column")
    title = args.title or value_column or args.symbol_column or ""
    rows.append(("title", f"{title}{'' if args.title else from_column}", "title"))
    legend = args.legend_title or value_column or ""
    if legend:
        rows.append(("legend",
                     f"{legend}{'' if args.legend_title else from_column}",
                     "legend_title"))
    symbol_legend = args.symbol_legend_title or args.symbol_column or ""
    if symbol_legend:
        rows.append(("symbol-legend",
                     f"{symbol_legend}{'' if args.symbol_legend_title else from_column}",
                     "symbol_legend_title"))

    rows += [
        ("scope", reach, "map_scope"),
        ("layout", wording.label("layout", args.layout), "layout"),
        ("language", wording.label("language", args.language), "language"),
        ("classes", (f"{wording.label('classification', args.classification)} — "
                       f"{wording.count('table.class-count', 'classes', bins['classes'])}"
                       if bins else unknown), "classification"),
        ("labels", wording.label("labels", args.labels), "labels"),
        ("repeated-rows", (wording.label("aggregate", method)
                      if method in wording.VALUES["aggregate"] else unknown), "aggregate"),
        ("output", (f"{args.formats.upper()} {args.dpi} dpi"
                    + ("" if args.no_html else messages.text("table.with-html"))), "formats"),
    ]

    numbered = []
    for number, (name, value, setting) in enumerate(rows, 1):
        row: dict[str, Any] = {"number": number, "item": wording.field(name),
                               "value": str(value)}
        if setting and setting not in chosen:
            row["note"] = auto
        if setting in wording.ALWAYS_SAFE or setting in among:
            offer = wording.menu(setting, _current(args, setting, scope, method),
                                 among=among.get(setting))
            if offer:
                row.update(question=offer["question"], choices=offer["choices"])
        numbered.append(row)

    settings = {wording.field(name): str(value) for name, value, _ in rows}
    must_ask = [wording.ask(setting, getattr(args, setting, None))
                for setting in THEIRS_TO_CHOOSE if setting not in chosen]
    # SKILL.md has asked for a title since its first version, and a real Codex
    # run drew three maps without asking anything at all. Saying so once more
    # would be the same instruction that already failed; withholding the code
    # is the thing that works.
    if "title" not in chosen:
        must_ask.append(wording.ask_in_words(
            "title",
            _title_ingredients(args, joined, value_column, scope, prepared,
                               country_name)))
    return settings, numbered, must_ask


def _current(args, setting: str, scope: str, method: str) -> str | None:
    """What a setting resolved to, which is not always what was typed.

    ``--map-scope auto`` becomes a real framing and ``--aggregate auto`` becomes
    a real method. The menu has to mark the resolved value as the current one,
    or the reader is offered, as an alternative, the thing they already have.
    """
    if setting == "map_scope":
        return scope
    if setting == "aggregate":
        return method
    return getattr(args, setting, None)


def _period_pins(args) -> list[str]:
    """The caller's own filters that pin the period column, if any.

    A column with four quarters in the workbook can arrive at the animation
    with one, because a ``--where`` pinned it. Saying only "this column holds
    one period" then reads as a fault in the workbook, and sends the reader
    looking through a spreadsheet for something that is not wrong with it.
    """
    if not args.period_column:
        return []
    return [f"{column}={value}"
            for column, value in longform.parse_where(args.where or [])
            if column == args.period_column]


def _map_label(args, value_column, ctx) -> str:
    """What the interactive page's map picker shows.

    A province name comes from the shapefile and stays Vietnamese in any
    language; "national" is the script's own word, so it gets translated.
    """
    lang = i18n.normalise(args.language)
    where = i18n.t(lang, "scope_national") if ctx["name"] == "national" else ctx["name"]
    return f"{args.title or value_column or ''} — {where}".strip(" —")


#: Fallback when a map has neither a title nor a value column to be named
#: after. ``render`` now refuses to issue a code without a title, so this is
#: close to unreachable — but it is a *file name*, and a file name written in
#: one language while the rest of the folder is written in another is the kind
#: of thing nobody notices until they are sorting a hundred of them.
UNTITLED = "map"


def map_basename(args, value_column, ctx) -> tuple[str, str]:
    """The stem of every file this map writes: ``(family, base)``.

    Three kinds of word meet here and they follow three different rules.

    * The **title** is the user's, in whatever language they wrote it.
    * The **place** is ``ctx["name"]`` — a province name comes from the
      boundary file and stays as it is, while ``national`` is the engine's own
      word for "all of it".
    * The **layout**, the **kind of file** (``_data``, ``_metadata``) and the
      word ``national`` are the engine's own structural markers, and they are
      English in every language, exactly like ``run_manifest.json`` beside
      them. Translating them would give the two editions of one map two
      different *structural* names, so a script looking for ``*_data.csv``
      would find the English run and miss the Vietnamese one.

    ``family`` groups the editions; ``base`` names one of them. The language
    suffix is what keeps a Vietnamese and an English edition of the same map
    side by side in one folder without either overwriting the other, and the
    layout is in the name because without it a second render of the same map in
    the other layout overwrites the first silently while ``run_manifest`` goes
    on listing both.
    """
    family = (dataio.slugify(f"{args.title or value_column or UNTITLED} {ctx['name']}")
              + f"_{args.layout}")
    return family, f"{family}_{i18n.suffix(args.language)}"


def _animation(deps, args, run_dir, joined, contexts, thematic, provinces_gdf,
               value_column, value_info, symbol_info, name_field, font_info,
               issues) -> dict[str, Any]:
    """The series as video, on one classification shared by every frame.

    Several map frames — one per province, say — each become their own video,
    but the class breaks and the circle scale are computed once across the whole
    set. That is the point: a colour has to mean the same thing in Nghệ An as in
    Hà Nội, and across every period of both, or the series cannot be compared
    with itself.
    """
    if not args.period_column:
        raise SystemExit(messages.text("error.animation-needs-period-column"))

    frames = period_utils.ordered(joined[args.period_column])
    if len(frames) < 2:
        # The column may hold plenty of periods and have been pinned to one by
        # the caller's own filter. Naming the column alone reads as a fault in
        # the workbook, and sends the reader looking for one that is not there.
        pinned = _period_pins(args)
        if pinned:
            whole = period_utils.ordered(
                dataio.read_table(deps, dataio.project_path(
                    Path(args.project_root).resolve(), args.excel),
                    args.sheet)[args.period_column])
            raise SystemExit(messages.text(
                "error.animation-period-pinned", column=args.period_column,
                filters=" and ".join(f"--where {p}" for p in pinned),
                count=len(whole)))
        raise SystemExit(messages.text("error.animation-needs-two-periods",
                                       singular=len(frames) == 1,
                                       column=args.period_column, count=len(frames)))
    unreadable = period_utils.unreadable(joined[args.period_column])
    if unreadable:
        issues.append(guardrails._issue(
            "unreadable-periods", guardrails.WARNING, counts=len(unreadable),
            fmt={"count": len(unreadable),
                 "periods": ", ".join(str(u) for u in unreadable[:5])}))

    values_by_period: dict[Any, dict[int, float]] = {}
    symbols_by_period: dict[Any, dict[int, float]] | None = ({} if args.symbol_column else None)
    for period in frames:
        sub = joined[joined[args.period_column].astype(str) == str(period)]
        if sub.empty:
            values_by_period[period] = {}
            if symbols_by_period is not None:
                symbols_by_period[period] = {}
            continue
        values_by_period[period] = aggregate.combine(
            deps, sub, "__shape_id", value_column, value_info, args.aggregate).to_dict()
        if symbols_by_period is not None:
            symbols_by_period[period] = aggregate.combine(
                deps, sub, "__shape_id", args.symbol_column, symbol_info, "auto").to_dict()

    pooled = [v for table in values_by_period.values() for v in table.values()
              if v is not None and not (isinstance(v, float) and v != v)]
    if not pooled:
        raise SystemExit(messages.text("error.no-values-after-periods"))
    bins = classify.compute_bins(pooled, args.classification, args.classes, value_info,
                                 center_zero=(args.map_type == "change"))
    bins["notes"].append(messages.text("read.shared-classes-over-time"))
    issues += guardrails.check_classes(bins, len(pooled))

    symbol_scale: dict[str, float] = {}
    if symbols_by_period is not None:
        symbol_scale = classify.symbol_scale(
            [v for table in symbols_by_period.values() for v in table.values()
             if v is not None and not (isinstance(v, float) and v != v)])

    wanted = args.animation_formats
    shared: dict[str, Any] = {"period": [str(p) for p in frames],
                              "shared_bins": bins,
                              "shared_symbol_scale": symbol_scale}
    made: list[dict[str, Any]] = []
    for ctx in contexts:
        ids = set(ctx["frame"]["__shape_id"])
        frame = thematic[thematic["__shape_id"].isin(ids)].copy()
        # coverage is counted against this frame, not the whole series: a
        # province's video should not claim units that belong to another one
        with_data = len({sid for table in values_by_period.values()
                         for sid in table} & ids)
        spec = _build_spec(args, ctx, value_column, value_info, symbol_info, bins,
                           symbol_scale, name_field,
                           aggregate.resolve(args.aggregate, value_info),
                           with_data, frame)
        spec["labels"] = "off"
        spec["diverging"] = args.map_type == "change"
        _, base = map_basename(args, value_column, ctx)

        common = dict(frame=frame, periods=frames, values_by_period=values_by_period,
                      symbols_by_period=symbols_by_period, spec=spec, fonts=font_info,
                      provinces=provinces_gdf, locator_name=ctx["locator"],
                      out_dir=run_dir, name=base)
        one: dict[str, Any] = {"map_name": ctx["name"], "units_with_data": with_data}
        if wanted in ("video", "both"):
            one["video"] = animate.build(deps, **common)
        if wanted in ("html", "both"):
            one["html"] = interactive.build(
                deps, label=_map_label(args, value_column, ctx), **common)
        dataio.write_json(run_dir / f"{base}_metadata.json",
                          {**shared, **one, "arguments": vars(args)})
        made.append(one)

    if len(made) == 1:
        return {**shared, **made[0]}
    # one entry per map frame, plus how many there are. "khung" survived the
    # rename because it carries no diacritic and the accent scan looked for one.
    return {**shared, "maps": made, "map_count": len(made)}


def _settle_language(args: argparse.Namespace) -> None:
    """Give ``--language`` a real value, once, at the entrance to drawing.

    The flag carries no argparse default so the gate can tell "the user chose
    Vietnamese" from "nobody asked" — ``chosen_explicitly`` answers that, and
    everything downstream wants an actual language. Filling it in here rather
    than in ``main`` keeps the two facts in one place: whoever reaches a render
    gets the resolved value, whichever way they arrived.
    """
    if getattr(args, "language", None) is None:
        args.language = i18n.DEFAULT


def _needs_place_column(args) -> bool:
    """Whether this run has to be told which column holds the place names.

    Written as a function rather than a condition inside the guard because the
    first version of that guard was a condition, and it blocked the one case
    that legitimately has no place column: a point map placed from coordinates.
    Every test stayed green while it did, which is the argument for naming the
    rule and testing it directly.

    The exemption is not a new idea — ``command_render`` already calls that case
    ``coordinates_only``. This has to agree with it, or the two drift apart.
    """
    if args.map_type == "point" and not args.province_column             and not args.commune_column:
        return False                      # coordinates_only: no matching at all
    if args.province_column:
        return False
    return args.admin_level == "province" or not args.commune_column


def command_render(args: argparse.Namespace) -> None:
    _settle_language(args)
    if getattr(args, "layer", None):
        return _render_layers(args)
    root = Path(args.project_root).resolve()
    deps = dataio.load(require_geo=True, require_plot=True)
    font_info = fonts.install(deps.matplotlib)

    excel = dataio.project_path(root, args.excel)
    # Both of these used to reach pandas and come back as a traceback: a missing
    # workbook as FileNotFoundError, a missing place column as KeyError: None.
    # A person reading either one learns nothing about what to do next.
    if excel is None or not excel.exists():
        raise SystemExit(messages.text("error.workbook-not-found", file=args.excel))
    if _needs_place_column(args):
        raise SystemExit(messages.text("error.place-column-required",
                                       flag="--province-column", level="province"))
    df = dataio.read_table(deps, excel, args.sheet)
    sheets = dataio.read_sheets(deps, excel)
    dictionary = dataio.read_data_dictionary(deps, excel, sheets)

    # Slice before anything else: the match review, the aggregation and every
    # warning should describe the rows actually being drawn, not the whole sheet.
    df, slice_note = _apply_where(df, args.where)

    country = getattr(args, "country", None)
    tier = dataio.resolve_tier(root, args.admin_level, country=country)
    admin_level = tier["role"]
    # Set before a single string is looked up, so nothing is drawn in one
    # language and labelled in another.
    i18n.use(_map_text(getattr(args, "map_text", None)))
    # carried on args so the plate builder, several calls down, can name the
    # tier the way the country does without another lookup
    args.tier_folder = tier["folder"]
    boundary_notes: list[dict[str, Any]] = []
    shapes = dataio.load_shapes(deps, root, admin_level, notes=boundary_notes,
                                country=country)
    fields = dataio.shape_fields(shapes, admin_level)
    # Resolved once, from the province tier, and used for every frame on the
    # page: the map and the locator beside it have to agree.
    thematic_crs = dataio.run_thematic_crs(deps, root, country=country)
    name_field = fields["commune"] if admin_level == "commune" else fields["province"]

    # --- match ------------------------------------------------------------
    # a coordinate-only point map needs no name matching at all
    coordinates_only = (args.map_type == "point" and not args.province_column
                        and not args.commune_column)
    review_path = dataio.project_path(root, args.match_review) if args.match_review else None
    if review_path is None and args.run_folder:
        # reuse exactly the table the user reviewed during profiling
        candidate = (root / "output" / args.run_folder
                     / review_filename(admin_level, excel, args.sheet))
        if candidate.exists():
            review_path = candidate

    # Read once from the profile, exactly as `profile` does: the index and
    # every lookup into it have to strip the same words or nothing matches.
    country_reading = dataio.read_country(
        deps, dataio.shapefile_root(root), country)
    affixes = dataio.name_affixes(country_reading)
    p_alias = dataio.alias_column(country_reading, dataio.COARSE)
    c_alias = dataio.alias_column(country_reading, dataio.FINE)

    review: list[dict[str, Any]] = []
    if coordinates_only:
        review = []
    elif review_path and review_path.exists():
        review = dataio.read_csv(review_path)
        if not _review_is_usable(review, admin_level):
            review = []          # belongs to another dataset; recompute below
    if not review and not coordinates_only:
        province_shapes = dataio.load_shapes(deps, root, dataio.COARSE,
                                             country=country)
        p_fields = dataio.shape_fields(province_shapes, dataio.COARSE)
        index = _province_index(province_shapes, p_fields["province"], affixes,
                                p_alias)
        if admin_level == "province":
            review = matching.review_province(
                [{"province": v} for v in df[args.province_column].tolist()], index)
        else:
            by_province = _commune_index_by_province(shapes, fields, affixes, c_alias)
            rows = [{"province": (df[args.province_column].iloc[i]
                                  if args.province_column else None),
                     "commune": df[args.commune_column].iloc[i]} for i in range(len(df))]
            review = matching.review_commune(rows, index, by_province,
                                             prefs.recall_overrides(root, "commune"))
    match_summary = matching.summarize(review)

    # --- attach shape ids -------------------------------------------------
    drop_ambiguous = args.ambiguous == "drop"
    match_summary["ambiguous_dropped"] = drop_ambiguous
    lookup = matching.shape_lookup(review, admin_level, drop_ambiguous=drop_ambiguous)

    def row_key(i: int) -> str:
        if admin_level == "province":
            return str(df[args.province_column].iloc[i]).strip()
        p = str(df[args.province_column].iloc[i]).strip() if args.province_column else ""
        return f"{p}|{str(df[args.commune_column].iloc[i]).strip()}"

    df = df.copy()
    if args.period_column and args.period and not args.animate:
        df = df[df[args.period_column].astype(str) == str(args.period)].reset_index(drop=True)
    if coordinates_only:
        df["__shape_id"] = None
        joined = df.copy()
    else:
        df["__shape_id"] = [lookup.get(row_key(i)) for i in range(len(df))]
        joined = df[df["__shape_id"].notna()].copy()
        joined["__shape_id"] = joined["__shape_id"].astype(int)
        if joined.empty:
            raise SystemExit(messages.text("error.no-rows-matched", file="match_review.csv"))

    # --- semantics + aggregation -----------------------------------------
    columns = profiling.describe_columns(deps, df, dictionary)
    weight_series = {c["column"]: df[c["column"]].tolist() for c in columns
                     if c["semantic"] in {sem.COUNT, sem.PERCENT, sem.RATE_PER, sem.POINT}}
    for info in columns:
        if info["semantic"] in {sem.PERCENT, sem.RATE_PER, sem.POINT}:
            info["weight_column"] = sem.find_denominator(info["column"], columns,
                                                        weight_series)
    by_name = {c["column"]: c for c in columns}

    value_column = args.value_column or args.category_column
    ratio_note = None
    if args.numerator or args.denominator or args.fill_indicator or args.symbol_indicator:
        value_column, joined, ratio_note = _build_long_columns(args, joined, by_name)
    else:
        # A wide sheet can ask for a quotient too. The long sheet has had this
        # since the beginning; the wide one refused it and the user had to add a
        # rate column to their workbook by hand.
        made = build_wide_ratio(joined, by_name, value_column)
        if made:
            value_column = made
        made = build_wide_ratio(joined, by_name, args.symbol_column)
        if made:
            args.symbol_column = made
    if args.map_type == "change":
        if not (args.baseline_column and args.comparison_column):
            raise SystemExit(messages.text("error.change-needs-two-columns"))
        value_column = i18n.t(args.language, "change_title",
                              comparison=args.comparison_column,
                              baseline=args.baseline_column)
        joined[value_column] = joined[args.comparison_column] - joined[args.baseline_column]
        by_name[value_column] = sem._pack(sem.POINT, value_column, "percentage point")
    if args.map_type == "graduated-symbol" and not args.symbol_column:
        raise SystemExit(messages.text("error.graduated-needs-symbol-column"))
    # a proportional-symbol map needs only the symbol column; area fills need a value
    if not value_column and args.map_type not in ("boundary", "point", "graduated-symbol"):
        raise SystemExit(messages.text("error.needs-value-column"))

    # --- coordinates for point maps --------------------------------------
    points = None
    if args.map_type == "point":
        lon_col, lat_col = args.lon_column, args.lat_column
        if not (lon_col and lat_col):
            coords = profiling.coordinate_candidates(columns)
            lon_col = lon_col or coords["lon"]
            lat_col = lat_col or coords["lat"]
        if not (lon_col and lat_col):
            raise SystemExit(messages.text("error.point-needs-coordinates"))
        valid = joined[joined[lon_col].notna() & joined[lat_col].notna()]
        if valid.empty:
            raise SystemExit(messages.text("error.no-rows-with-coordinates",
                                           lon=lon_col, lat=lat_col))
        located = deps.gpd.GeoSeries(
            deps.gpd.points_from_xy(valid[lon_col], valid[lat_col]), crs="EPSG:4326"
        ).to_crs(thematic_crs)
        points = {"x": located.x.tolist(), "y": located.y.tolist(),
                  "skipped_missing_coords": int(len(joined) - len(valid))}
        _dress_points(points, valid, args, by_name)

    value_info = by_name.get(value_column, {"semantic": sem.UNKNOWN, "column": value_column})
    duplicates = aggregate.duplicate_count(joined, "__shape_id") if not coordinates_only else 0

    issues: list[dict[str, Any]] = list(guardrails.check_matching(match_summary))
    country_profile = dataio.read_country(
        deps, dataio.shapefile_root(root), tier["country"])
    detached = country_profile.get("detached_land")
    # Read from the profile, not decided here. Vietnam's 111°E used to be a
    # constant inside the drawing code, which meant no other country could ever
    # have an inset and none was ever told why.
    inset_lon = insets.meridian(country_profile)
    inset_label = insets.inset_label(country_profile)
    values = None
    method = "n/a"
    if value_column and args.map_type != "boundary" and not coordinates_only:
        method = aggregate.resolve(args.aggregate, value_info)
        issues += guardrails.check_aggregation(value_info, method, duplicates)
        issues += guardrails.check_colour_choice(value_info, args.map_type)
        values = aggregate.combine(deps, joined, "__shape_id", value_column, value_info,
                                   args.aggregate)
        issues += guardrails.check_percent_range(values.tolist(), value_info)
        issues += guardrails.check_diverging(values.tolist(), value_info)

    symbol_info = by_name.get(args.symbol_column, {}) if args.symbol_column else {}
    symbols = None
    if args.symbol_column:
        symbols = aggregate.combine(deps, joined, "__shape_id", args.symbol_column,
                                    symbol_info, "auto")

    # --- what gets drawn --------------------------------------------------
    scope, contexts = _contexts(args, shapes, fields, review, admin_level)
    thematic = dataio.to_thematic_crs(shapes, thematic_crs)
    provinces_gdf = dataio.to_thematic_crs(
        dataio.load_shapes(deps, root, dataio.COARSE, country=country), thematic_crs)

    prepared = []
    for ctx in contexts:
        keep = set(ctx["frame"]["__shape_id"])
        frame = thematic[thematic["__shape_id"].isin(keep)].copy()
        frame["__value"] = frame["__shape_id"].map(values) if values is not None else None
        if symbols is not None:
            frame["__symbol"] = frame["__shape_id"].map(symbols)
        # carried on the context rather than passed down: both the still and the
        # animation build their spec from a context, and a meridian that reached
        # only one of them would frame the two editions of the same map
        # differently
        prepared.append({**ctx, "frame": frame,
                         "inset_meridian": inset_lon,
                         "inset_label": inset_label})

    # one classification and one symbol scale for the whole job, so the same
    # colour and the same circle size mean the same thing on every sheet
    bins = None
    if values is not None and args.map_type in {"choropleth", "choropleth-symbol", "change"}:
        groups = {c["name"]: c["frame"]["__value"].dropna().tolist() for c in prepared}
        pooled = [v for vals in groups.values() for v in vals]
        if not pooled:
            raise SystemExit(messages.text("error.no-values-after-matching"))
        center_zero = args.map_type == "change"
        bins = (classify.shared_bins(groups, args.classification, args.classes,
                                     value_info, center_zero)
                if len(prepared) > 1
                else classify.compute_bins(pooled, args.classification, args.classes,
                                           value_info, center_zero))
        issues += guardrails.check_classes(bins, len(pooled))

    symbol_scale: dict[str, float] = {}
    if symbols is not None:
        pooled = [v for c in prepared for v in c["frame"]["__symbol"].dropna().tolist()]
        symbol_scale = classify.symbol_scale(pooled)

    # --- the gate ---------------------------------------------------------
    # Everything above decided what the map will be; nothing above drew it. The
    # plan goes to the person first, and only the code derived from it unlocks
    # the drawing. Placed before the run folder is opened, so a plan that is
    # never agreed leaves nothing behind.
    # Before any code is issued: a plate whose words cannot be drawn is worse
    # than one that is refused, because it looks finished. The rule has always
    # been that the run stops rather than substituting a typeface — it just did
    # not cover the font that loads and has no glyph for the text.
    lettered = [args.title, args.legend_title, args.symbol_legend_title,
                args.subtitle, args.insight, args.source_note, args.footnote,
                value_column, args.symbol_column]
    lettered += list(i18n.overrides().values())
    for context in prepared:
        frame = context["frame"]
        if name_field in frame.columns:
            lettered += [str(v) for v in frame[name_field]]
    absent = fonts.undrawable(lettered)
    if absent:
        raise SystemExit(messages.text("error.font-cannot-draw",
                                       count=len(absent),
                                       characters=" ".join(absent[:12])))

    settings, numbered, must_ask = _plan(args, excel, joined, value_column,
                                         scope, prepared, method, bins,
                                         country_reading.get('country_name'))
    # No code unlocks a plan that still has a question open in it. The hash
    # covers the settings, and a defaulted language reads the same in the table
    # as a chosen one — so without this an agent could take the code from its own
    # planning run, never ask, and draw the default it invented. The settings the
    # person owns have to arrive on the command line, and they only can once
    # somebody has been asked.
    if must_ask or not confirm.matches(args.confirmed, settings):
        emit(confirm.gate(settings, numbered,
                          guardrails.summarize(issues)["items"], must_ask,
                          _command_line(),
                          language_stated="messages" in getattr(
                              args, "chosen_explicitly", set()),
                          # what to offer for the text on the map, and why —
                          # the machine's own setting and the country of the
                          # boundaries, reported separately because they
                          # disagree often enough to matter
                          language_hint=i18n.suggest(detect.country_language(
                              country_profile.get("country_name")))))
        return

    # --- draw -------------------------------------------------------------
    run_dir = dataio.create_run_dir(root, args.run_folder)

    if args.animate:
        series = _animation(deps, args, run_dir, joined, prepared, thematic, provinces_gdf,
                            value_column, value_info, symbol_info, name_field, font_info,
                            issues)
        emit({"output_folder": str(run_dir), "over_time": series,
              "open_files": dataio.openable(run_dir),
              "warnings": guardrails.summarize(issues)["items"]})
        return

    outputs, per_map = [], []
    for ctx in prepared:
        frame = ctx["frame"]
        # Two different questions, and they used to share one answer. Coverage is
        # about the fill: grey area reads as "surveyed and found nothing". What
        # the report calls units_with_data is about the map as a whole, and
        # counting only the fill made a proportional-symbol map drawing fourteen
        # circles report zero — while its own subtitle said 14/126.
        filled = int(frame["__value"].notna().sum()) if values is not None else 0
        with_data = _units_shown(frame, values is not None, symbols is not None)
        issues_ctx = list(issues)
        # coverage only matters when areas are filled by colour
        if args.map_type in {"choropleth", "choropleth-symbol", "change", "categorized"}:
            issues_ctx += guardrails.check_coverage(filled, len(frame))

        spec = _build_spec(args, ctx, value_column, value_info, symbol_info,
                           bins, symbol_scale, name_field, method, with_data, frame,
                           points_count=len(points["x"]) if points else None)
        spec["points"] = points
        result = render.draw(deps, frame=frame, spec=spec, fonts=font_info,
                             provinces=provinces_gdf, locator_name=ctx["locator"])
        if symbol_scale:
            span_y = frame.total_bounds[3] - frame.total_bounds[1]
            radii = render.symbol_radii(frame["__symbol"].dropna().tolist(),
                                        symbol_scale, span_y)
            issues_ctx += guardrails.check_symbol_occlusion(
                max(radii) if radii else 0.0, render.median_feature_width(frame))

        family, base = map_basename(args, value_column, ctx)
        # the page needs the live axes, so capture before save() closes the figure
        if not args.no_html:
            webpage.stash(run_dir, base, webpage.capture_still(
                result["plate"], frame, spec, family=family,
                label=_map_label(args, value_column, ctx),
                fills=result["fills"],
                point_counts=_points_per_unit(deps, frame, points)))
        written = render.save(result["plate"], run_dir, base, args.formats, dpi=args.dpi)
        outputs.extend(str(p) for p in written)
        table = _drawn_table(run_dir, base, frame, name_field, value_column,
                             value_info, symbol_info, args, bins)
        # Whether an inset was drawn is only known once the map is drawn, and
        # the warning is about the two together: land far from the main body,
        # and nothing done about it.
        if scope == "national":
            issues_ctx = issues_ctx + guardrails.check_detached_territory(
                detached, result.get("inset") is not None, lang=args.messages)
        meta = {"map_name": ctx["name"], "files": [str(p) for p in written],
                "data_table": table,
                "units_with_data": with_data, "units_in_frame": int(len(frame)),
                "labels": result["labels"], "inset": result.get("inset"),
                "overflow": result.get("overflow") or [],
                "detached_land": detached,
                "warnings": guardrails.summarize(issues_ctx)}
        dataio.write_json(run_dir / f"{base}_metadata.json", {**meta, "arguments": vars(args)})
        per_map.append(meta)

    if review:
        dataio.write_csv(run_dir / review_filename(admin_level, excel, args.sheet), review)
    prefs.remember_choices(root, excel, args.sheet, {
        "admin_level": admin_level, "map_type": args.map_type, "layout": args.layout,
        "value_column": args.value_column, "symbol_column": args.symbol_column,
        "province_column": args.province_column, "commune_column": args.commune_column,
        "classification": args.classification, "classes": args.classes,
        "labels": args.labels, "map_scope": scope, "language": args.language,
    })

    job = {
        "scope": scope, "maps": per_map,
        "language": i18n.normalise(args.language), "layout": args.layout,
        "shared_bins": bins, "shared_symbol_scale": symbol_scale,
        "aggregation": aggregate.describe(args.aggregate, value_info,
                                            value_info.get("weight_column"),
                                            args.language),
        "name_matching": match_summary, "font": font_info, "arguments": vars(args),
        # what slice of a long sheet this map actually drew; without it a PNG
        # from a 70.000-row export cannot be traced back to its rows
        "data_slice": slice_note, "ratio": ratio_note,
    }
    # one request may render several times (two languages, two layouts); the
    # manifest keeps every job instead of the last one overwriting the rest
    _append_manifest(run_dir, "renders", job)

    # rebuilt from every capture in this request folder, so a second render —
    # the English edition, another layout — extends the page instead of
    # replacing what the first one produced
    page = None if args.no_html else webpage.build(run_dir, webpage.STILL)

    emit({"output_folder": str(run_dir), "image_files": outputs, "interactive_page": page,
          # ready-made addresses, so nothing is left for the agent to construct
          "open_files": dataio.openable(run_dir),
          **({"codepage_repair": boundary_notes} if boundary_notes else {}),
          "projection": thematic_crs,
          "warnings": guardrails.summarize(issues)["items"], "maps": per_map})


def _build_spec(args, ctx, value_column, value_info, symbol_info, bins,
                symbol_scale, name_field, method, with_data, frame,
                points_count: int | None = None) -> dict[str, Any]:
    lang = i18n.normalise(args.language)
    kicker = args.subtitle or i18n.kicker(lang, args.admin_level,
                                          tier=getattr(args, "tier_folder", None))
    if ctx["locator"] and ctx["locator"].upper() not in kicker.upper():
        kicker = f"{kicker}  ·  {ctx['locator'].upper()}"

    title = args.title or (value_column or args.symbol_column or "")
    # a proportional-symbol map describes itself through the symbol column
    if value_column:
        insight_col, insight_info = "__value", value_info
    elif args.symbol_column:
        insight_col, insight_info = "__symbol", symbol_info
    else:
        insight_col, insight_info = None, {}
    insight = args.insight or _auto_insight(frame, insight_col, insight_info,
                                            name_field, lang, points_count,
                                            decimals=classify.label_decimals(bins))
    source = args.source_note or i18n.t(lang, "source", file=Path(args.excel).name)
    fills_areas = args.map_type in {"choropleth", "choropleth-symbol", "change", "categorized"}
    method_note = args.footnote or _auto_method(args, bins, method, fills_areas, lang)
    # the side column also hosts the locator, so keep it whenever one is drawn;
    # otherwise its caption would sit alone in the margin
    wants_locator = bool(ctx["locator"]) and args.locator != "off"

    return {
        "language": lang,
        "side_panel": bool(fills_areas or args.symbol_column or wants_locator),
        "layout": args.layout, "map_type": args.map_type,
        "value_column": "__value" if value_column else None,
        "value_info": value_info,
        "symbol_column": "__symbol" if args.symbol_column else None,
        "symbol_info": symbol_info,
        "bins": bins, "symbol_scale": symbol_scale,
        "name_field": name_field,
        "labels": args.labels, "label_fontsize": args.label_fontsize,
        "legend_title": (args.legend_title
                         or (i18n.t(lang, "change_legend") if args.map_type == "change"
                             else value_column or "")),
        "symbol_legend_title": args.symbol_legend_title or args.symbol_column or "",
        "kicker": kicker, "title": title, "insight": insight,
        "source": source, "method": method_note,
        "locator": args.locator != "off",
        "province_name_field": "ten_tinh",
        "inset_meridian": ctx.get("inset_meridian"),
        "inset_label": ctx.get("inset_label"),
        "dpi": args.dpi,
    }


def _auto_insight(frame, column, info, name_field, lang: str | None = None,
                  points_count: int | None = None,
                  decimals: int | None = None) -> str:
    """One descriptive sentence, taken only from what the map actually shows.

    ``decimals`` comes from the class breaks, so the number in this sentence and
    the number on the unit below it are the same number written the same way. A
    map whose subtitle read "highest at 0%" over a label reading "0.019%" is two
    statements about one figure, and the reader has no way to tell which to
    believe.
    """
    total = len(frame)
    if points_count is not None:
        return i18n.t(lang, "insight_points", points=points_count, total=total)
    if not column or column not in frame.columns:
        return i18n.t(lang, "insight_frame", total=total)
    data = frame[frame[column].notna()]
    with_data = len(data)
    if not with_data:
        return i18n.t(lang, "insight_frame", total=total)

    if info.get("semantic") == sem.CATEGORY or data[column].dtype == object:
        counts = data[column].astype(str).value_counts()
        return i18n.t(lang, "insight_category", with_data=with_data, total=total,
                      name=counts.index[0], count=int(counts.iloc[0]))
    try:
        top = data.loc[data[column].idxmax()]
    except (TypeError, ValueError):
        return i18n.t(lang, "insight_plain", with_data=with_data, total=total)
    return i18n.t(lang, "insight_values", with_data=with_data, total=total,
                  name=top[name_field],
                  value=sem.format_value(top[column], info, decimals=decimals,
                                         lang=lang))


def _auto_method(args, bins, method, fills_areas: bool, lang: str | None = None) -> str:
    parts = []
    if bins:
        name = i18n.t(lang, f"class_{args.classification}")
        if name == f"class_{args.classification}":
            name = args.classification
        parts.append(i18n.t(lang, "method_classes", classes=bins["classes"], method=name))
    if args.symbol_column:
        parts.append(i18n.t(lang, "method_symbol"))
    if method not in ("n/a", "sum"):
        label = i18n.t(lang, f"agg_{method}")
        parts.append(i18n.t(lang, "method_aggregate",
                            method=label if label != f"agg_{method}" else method))
    if args.map_type == "point":
        parts.append(i18n.t(lang, "method_points"))
    if fills_areas:
        parts.append(i18n.t(lang, "method_grey"))
    return " ".join(parts)


# --------------------------------------------------------------------------
def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="easy-map engine")
    parser.add_argument("--project-root", default=".",
                        help="also accepted after the subcommand, e.g. list --project-root .")
    parser.add_argument("--messages", default=messages.DEFAULT,
                        choices=list(messages.LANGUAGES),
                        help="language of the warnings and reasons returned to the agent — "
                             "the language of the conversation. Not --language, "
                             "which is the language lettered on the map")
    sub = parser.add_subparsers(dest="command", required=True)

    sr = sub.add_parser("start-run",
                        help="open a folder for one request and print its name")
    sr.add_argument("--run-folder",
                    help="name it by hand; omit for the current timestamp")

    sub.add_parser("list", help="list the workbooks, sheets and boundary files")

    im = sub.add_parser("import",
                        help="put a file the user attached into input/ and survey it")
    im.add_argument("--file", required=True,
                    help="path to the file the user attached to the conversation")
    im.add_argument("--run-folder",
                    help="record the import in this request's manifest; omit to use "
                         "the open start-run folder, if there is one")

    sv = sub.add_parser("survey",
                        help="which sheets in a workbook can be mapped; samples, so it is fast")
    sv.add_argument("--country", help="country folder under shapefiles/; omit when there is "
                                      "only one country")
    sv.add_argument("--excel",
                    help="omit to survey every workbook in input/")

    p = sub.add_parser("profile", help="read the dataset and propose a map")
    p.add_argument("--country", help="country folder under shapefiles/; omit when there is "
                                     "only one country")
    p.add_argument("--excel", required=True)
    p.add_argument("--sheet")
    p.add_argument("--admin-level", default="auto",
                   help="a role (province/commune) or a tier folder name, e.g. state")
    p.add_argument("--province-column")
    p.add_argument("--commune-column")
    p.add_argument("--run-folder",
                   help="folder for this request; omit to reuse the open "
                        "start-run folder")

    f = sub.add_parser("fix-match", help="remember a name match the user confirmed")
    f.add_argument("--country")
    f.add_argument("--admin-level", default="commune",
                   help="a role, or a tier folder name")
    f.add_argument("--province")
    f.add_argument("--name", required=True)
    f.add_argument("--shape-id", required=True)

    r = sub.add_parser("render", help="draw the map and write it out")
    r.add_argument("--country", help="country folder under shapefiles/; omit when there is "
                                     "only one country")
    r.add_argument("--excel", required=True)
    r.add_argument("--sheet")
    r.add_argument("--admin-level", required=True,
                   help="a role (province/commune) or a tier folder name, e.g. state")
    r.add_argument("--province-column")
    r.add_argument("--commune-column")
    r.add_argument("--match-review")
    r.add_argument("--map-type", default="choropleth",
                   choices=["choropleth", "choropleth-symbol", "graduated-symbol",
                            "categorized", "boundary", "change", "point"])
    r.add_argument("--value-column")
    r.add_argument("--symbol-column")
    r.add_argument("--category-column")
    r.add_argument("--baseline-column")
    r.add_argument("--comparison-column")
    r.add_argument("--period-column")
    r.add_argument("--period")
    r.add_argument("--animate", action="store_true",
                   help="build a map over time instead of a still; needs --period-column")
    r.add_argument("--animation-formats", default="both", choices=["video", "html", "both"],
                   help="an MP4/GIF video, an interactive HTML page, or both")
    r.add_argument("--lon-column", help="longitude column, for a point map")
    r.add_argument("--lat-column", help="latitude column, for a point map")
    r.add_argument("--point-color-column", metavar="COLUMN",
                   help="point map: the categorical column that decides dot colour "
                        "(kind of facility, priority level)")
    r.add_argument("--point-size-column", metavar="COLUMN",
                   help="point map: the numeric column that decides dot size; area is "
                        "proportional to the value, on the legend's own scale")
    r.add_argument("--aggregate", default="auto", choices=list(aggregate.METHODS))
    r.add_argument("--map-text", action="append", metavar="KEY=VALUE",
                   help="replace one string the engine letters on the map, e.g. "
                        "--map-text no_data='Aucune donnée'. Repeatable. Use it "
                        "to print a map in a Latin-script language other than "
                        "vi/en; the packaged fonts hold no Chinese, Cyrillic, "
                        "Thai or Lao, and the run stops rather than draw boxes")
    r.add_argument("--map-scope", default="auto",
                   choices=["auto", "national", "single-province", "province-series",
                            "matched-only"])
    r.add_argument("--province-series-threshold", type=int, default=PROVINCE_SERIES_THRESHOLD)
    # default None, not "vi": the gate needs to tell "the user chose Vietnamese"
    # apart from "nobody asked", and those are different things to report
    r.add_argument("--language", default=None, choices=list(i18n.LANGUAGES),
                   help="language of the machine-written text on the map, and the file-name "
                        "suffix. NOT inferred from the conversation — ask the user")
    r.add_argument("--confirmed", metavar="CODE",
                   help="the code from a previous run of this same command, once the "
                        "user has read the settings table and agreed. Without it the "
                        "command returns the table and draws nothing")
    r.add_argument("--layout", default="report", choices=["report", "banner"])
    r.add_argument("--classification", default="quantile", choices=list(classify.METHODS))
    r.add_argument("--classes", type=int, default=5)
    r.add_argument("--labels", default="both", choices=["off", "names", "values", "both"])
    r.add_argument("--label-fontsize", type=float, default=8.0)
    r.add_argument("--formats", default="png", choices=["png", "svg", "both"])
    r.add_argument("--dpi", type=int, default=220)
    r.add_argument("--layer", action="append", metavar="VARIABLE",
                   help="a variable to show; repeatable. The skill assigns each to a "
                        "channel (area fill / circles) and moves the overflow to a "
                        "second plate. Wide table: a column name. Long table (with "
                        "--indicator-column): an indicator value, or "
                        "'NUMERATOR / DENOMINATOR' for a rate")
    r.add_argument("--indicator-column", metavar="COLUMN",
                   help="long table: the column holding indicator names. With it, --layer "
                        "and --fill-indicator/--symbol-indicator take indicator values "
                        "rather than column names")
    r.add_argument("--fill-indicator",
                   help="the indicator value the fill uses, when the fill is not a ratio")
    r.add_argument("--fill-where", action="append", metavar="COLUMN=VALUE",
                   help="the slice the fill indicator is taken from; repeatable. Use it "
                        "when two indicators on one plate need different pins on the "
                        "same column")
    r.add_argument("--symbol-where", action="append", metavar="COLUMN=VALUE",
                   help="the slice the circle indicator is taken from; repeatable")
    r.add_argument("--where", action="append", metavar="COLUMN=VALUE",
                   help="keep the rows that match; repeatable. Required on a long table "
                        "so nothing is counted twice")
    r.add_argument("--ratio-column",
                   help="the former name of --indicator-column, kept for commands already written")
    r.add_argument("--numerator", help="the indicator value to use as the numerator")
    r.add_argument("--symbol-indicator",
                   help="ratio mode: the indicator value the circles use, taken from a "
                        "different slice than the numerator and denominator")
    r.add_argument("--denominator", help="the indicator value to use as the denominator")
    r.add_argument("--ambiguous", default="drop", choices=["drop", "keep"],
                   help="a row whose name matches several communes: 'drop' leaves it off the "
                        "map (default), 'keep' draws it on the first candidate")
    r.add_argument("--no-html", action="store_true",
                   help="skip the interactive HTML page; every request gets one by default")
    r.add_argument("--locator", default="auto", choices=["auto", "off"])
    r.add_argument("--title")
    r.add_argument("--subtitle")
    r.add_argument("--insight")
    r.add_argument("--legend-title")
    r.add_argument("--symbol-legend-title",
                   help="heading of the circle legend; defaults to the column name")
    r.add_argument("--source-note")
    r.add_argument("--footnote")
    r.add_argument("--run-folder",
                   help="folder for this request; omit to reuse the open "
                        "start-run folder")

    # accept --project-root and --messages either before or after the subcommand
    for child in (sub.choices.values()):
        child.add_argument("--project-root", dest="project_root_sub", default=None)
        child.add_argument("--messages", dest="messages_sub", default=None,
                           choices=list(messages.LANGUAGES))
    return parser


#: Flags whose presence means "the person decided this", as opposed to the skill
#: falling back to a default. argparse throws that distinction away, so it is
#: recovered from the raw argument list before parsing.
def _explicit(argv: list[str]) -> set[str]:
    named = {a.split("=", 1)[0] for a in argv if a.startswith("--")}
    return {flag.lstrip("-").replace("-", "_") for flag in named}


def main(argv: list[str] | None = None) -> None:
    speak_utf8()
    raw = list(sys.argv[1:] if argv is None else argv)
    args = build_parser().parse_args(argv)
    args.chosen_explicitly = _explicit(raw)
    if getattr(args, "project_root_sub", None):
        args.project_root = args.project_root_sub
    if getattr(args, "messages_sub", None):
        args.messages = args.messages_sub
    # the language of the conversation, which is not the language of the map:
    # an English-speaking officer often needs a Vietnamese map for a Vietnamese
    # audience, so --messages and --language are set independently
    messages.use(args.messages)
    {"start-run": command_start_run, "list": command_list, "import": command_import,
     "survey": command_survey, "profile": command_profile,
     "render": command_render, "fix-match": command_fix_match}[args.command](args)


if __name__ == "__main__":
    main()
