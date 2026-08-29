"""One practice workbook per province for the September training.

**Every number here is invented.** Nothing was observed, nothing was reported by
anybody, and none of it may be used for a report or a programme decision.

Eleven provinces, one file each, at commune level: outreach, commodities,
testing and treatment. Two things make these different from the other fixtures
in ``input/``.

**The figures hold together.** A positivity rate really is its positives over
its tests; an ART coverage rate really is patients over people in care; needles
are only distributed where the file says people who inject drugs are reached.
A table whose rate does not match its own numerator and denominator teaches the
skill's denominator check nothing, because the check works by *reproducing the
rate row by row* and gives up when it cannot.

**No two files are laid out alike.** The header row differs in every workbook -
row 1 in one, row 10 in another - and some carry a guidance sheet while others
do not. That is the point: participants should see the skill find the table
wherever it sits, rather than infer that it needs a tidy sheet with column names
on the first row. A file that fails is worth more here than a file that flatters.

Two files carry a note two rows below the table, as real reports do. The reader
takes the blank row and the note as two more rows of data, and the skill reports
them as unmatched rather than drawing them - which is the right answer and worth
showing. If ``nghe_an`` and ``quang_tri`` warn about two unmatched rows, that is
the note, not a fault in the data.

    uv run --offline --with pandas --with openpyxl --with geopandas \
        python tools/generate_training_data.py

Writes into ``input/``, which is not tracked.
"""

from __future__ import annotations

import sys
from pathlib import Path

import geopandas as gpd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "skills" / "easy-map" / "scripts"))

OUT = ROOT / "input"

DISCLAIMER = ("So lieu gia lap phuc vu tap huan. Khong phai so lieu giam sat "
              "HIV that va khong duoc dung cho bao cao hay quyet dinh chuyen mon.")

TITLE = Font(bold=True, size=13)
SUB = Font(size=10, color="444444")
HEAD = Font(bold=True, size=10, color="FFFFFF")
NOTE = Font(italic=True, size=9, color="6B7780")
FILL = PatternFill("solid", fgColor="1F5C99")
WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)

#: The eleven provinces named in Phu luc 2 of the training circular, spelled as
#: the boundary file spells them. The circular writes "Khanh Hoa" and "Thanh
#: Hoa" with the diacritic on the a rather than the o; the engine matches either,
#: but a fixture should not need the matcher's help to find its own province.
PROVINCES = [
    ("Hà Nội", "ha_noi"),
    ("Hải Phòng", "hai_phong"),
    ("Thái Nguyên", "thai_nguyen"),
    ("Quảng Ninh", "quang_ninh"),
    ("Nghệ An", "nghe_an"),
    ("Đà Nẵng", "da_nang"),
    ("Khánh Hòa", "khanh_hoa"),
    ("Thanh Hóa", "thanh_hoa"),
    ("Quảng Trị", "quang_tri"),
    ("Hà Tĩnh", "ha_tinh"),
    ("Huế", "hue"),
]


def settled(name: str, low: float, high: float, salt: int = 0) -> float:
    """A fixed value for a name: same name, same number, every run.

    Not ``random`` with a seed. A seeded generator gives the same *sequence*, so
    inserting one commune shifts every figure after it and every workbook looks
    changed between runs.
    """
    h = 0
    for ch in f"{name}|{salt}":
        h = (h * 131 + ord(ch)) % 1_000_003
    return low + (high - low) * (h / 1_000_003)


def pick(name: str, options, salt: int = 0):
    h = 0
    for ch in f"{name}|{salt}":
        h = (h * 131 + ord(ch)) % 1_000_003
    return options[h % len(options)]


# --------------------------------------------------------------------------
def cascade(province: str, commune: str) -> dict:
    """One commune's figures, built so that every derived value is derivable.

    Each step is a share of the step above it, in the order a programme works:
    a population, an estimated key population inside it, the part of that reached
    by outreach, commodities handed to those reached, tests done, positives
    found, people in care, people on treatment. Nothing is drawn independently
    that ought to be a fraction of something else, which is why the two rates
    below are exactly their own numerator over their own denominator.
    """
    key = f"{province}/{commune}"

    population = int(settled(key, 3_000, 42_000, 1))
    key_population = max(12, int(population * settled(key, 0.003, 0.012, 2)))
    reached = int(key_population * settled(key, 0.45, 0.88, 3))

    condoms = int(reached * settled(key, 8, 30, 4))
    # Needle and syringe programmes do not run everywhere: about two communes
    # in three here. A column of small non-zero numbers everywhere would suggest
    # a service that exists in every commune, which is not how the programme is
    # organised.
    injects = settled(key, 0, 1, 5) > 0.34
    needles = int(reached * settled(key, 40, 130, 6)) if injects else 0

    tested = int(reached * settled(key, 0.55, 0.95, 7)) + int(
        population * settled(key, 0.004, 0.02, 8))
    positive = max(0, int(tested * settled(key, 0.002, 0.032, 9)))
    positivity = round(100 * positive / tested, 2) if tested else 0.0

    in_care = positive + int(settled(key, 18, 260, 10))
    on_art = int(in_care * settled(key, 0.78, 0.985, 11))
    art_coverage = round(100 * on_art / in_care, 2) if in_care else 0.0
    prep = int(key_population * settled(key, 0.02, 0.19, 12))

    return {
        "Xã/phường": commune,
        "Dân số 15-49 tuổi": population,
        "Ước tính quần thể nguy cơ cao": key_population,
        "Số người được tiếp cận cộng đồng": reached,
        "Số bao cao su đã phát": condoms,
        "Số bơm kim tiêm đã phát": needles,
        "Số người được xét nghiệm HIV": tested,
        "Số ca dương tính": positive,
        "Tỷ lệ dương tính (%)": positivity,
        "Số người nhiễm HIV đang quản lý": in_care,
        "Số người đang điều trị ARV": on_art,
        "Tỷ lệ điều trị ARV (%)": art_coverage,
        "Số người đang dùng PrEP": prep,
        "Nhóm ưu tiên can thiệp": pick(key, ("Thấp", "Trung bình", "Cao",
                                             "Rất cao"), 13),
    }


def communes(frame, province: str) -> list[dict]:
    part = frame[frame["ten_tinh"].astype(str) == province]
    return [cascade(province, str(x)) for x in part["ten_xa"]]


# --------------------------------------------------------------------------
# the layouts
#
# Eleven of them, and no two the same. The header row is what matters most:
# participants should watch the skill find the column names on row 9 as readily
# as on row 1.

def sheet_of(book, name: str, rows: list[dict], columns: list[str],
             header_row: int, preamble: list[tuple[str, Font]] = (),
             footer: str | None = None, province: str | None = None):
    """Write one table, with its header on ``header_row``."""
    ws = book.create_sheet(name[:31])
    for offset, (text, font) in enumerate(preamble):
        cell = ws.cell(row=offset + 1, column=1, value=text)
        cell.font = font
    for index, column in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=index, value=column)
        cell.font = HEAD
        cell.fill = FILL
        cell.alignment = WRAP
    for r, row in enumerate(rows, start=header_row + 1):
        for c, column in enumerate(columns, start=1):
            value = (province if column in ("Tỉnh/thành phố", "Tỉnh")
                     else row.get(column))
            ws.cell(row=r, column=c, value=value)
    if footer:
        ws.cell(row=header_row + len(rows) + 2, column=1, value=footer).font = NOTE
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    for index, column in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(index)].width = min(
            34, max(12, len(column) * 0.95))
    return ws


def guidance_sheet(book, province: str, columns: list[str]):
    ws = book.create_sheet("Từ điển dữ liệu")
    ws.cell(row=1, column=1, value=f"Từ điển dữ liệu - {province}").font = TITLE
    ws.cell(row=2, column=1, value=DISCLAIMER).font = NOTE
    meaning = {
        "Xã/phường": ("Tên xã/phường theo bố cục hành chính 2025", "tên"),
        "Tỉnh/thành phố": ("Tên tỉnh/thành phố", "tên"),
        "Dân số 15-49 tuổi": ("Dân số trong độ tuổi 15-49", "người"),
        "Ước tính quần thể nguy cơ cao": ("Ước tính MSM, TCMT, PNBD", "người"),
        "Số người được tiếp cận cộng đồng": ("Tiếp cận viên gặp trực tiếp", "người"),
        "Số bao cao su đã phát": ("Bao cao su cấp phát trong kỳ", "chiếc"),
        "Số bơm kim tiêm đã phát": ("Bơm kim tiêm cấp phát; 0 = không có chương trình",
                                    "chiếc"),
        "Số người được xét nghiệm HIV": ("Lượt người xét nghiệm trong kỳ", "người"),
        "Số ca dương tính": ("Số ca có kết quả khẳng định dương tính", "ca"),
        "Tỷ lệ dương tính (%)": ("Số ca dương tính / số người xét nghiệm", "%"),
        "Số người nhiễm HIV đang quản lý": ("Số người nhiễm HIV được quản lý", "người"),
        "Số người đang điều trị ARV": ("Số người đang dùng thuốc ARV", "người"),
        "Tỷ lệ điều trị ARV (%)": ("Đang điều trị ARV / đang quản lý", "%"),
        "Số người đang dùng PrEP": ("Số người đang dùng PrEP", "người"),
        "Nhóm ưu tiên can thiệp": ("Xếp loại ưu tiên can thiệp", "nhóm"),
    }
    for index, head in enumerate(("Cột", "Ý nghĩa", "Đơn vị"), start=1):
        cell = ws.cell(row=4, column=index, value=head)
        cell.font = HEAD
        cell.fill = FILL
    for offset, column in enumerate(columns):
        what, unit = meaning.get(column, ("", ""))
        ws.cell(row=5 + offset, column=1, value=column)
        ws.cell(row=5 + offset, column=2, value=what)
        ws.cell(row=5 + offset, column=3, value=unit)
    for width, letter in ((36, "A"), (46, "B"), (12, "C")):
        ws.column_dimensions[letter].width = width


#: A commune column alone cannot be matched. Commune names repeat across the
#: country - 2,849 distinct over 3,321 - so ``matching.review_commune`` places a
#: commune *inside its province* and returns nothing when the province is
#: missing. The first draft of these workbooks left the province in the title
#: block only, and ten of the eleven were refused with "no rows matched".
#:
#: So every file carries the province, but not in the same place or under the
#: same heading: the engine finds the column by comparing its values against the
#: boundary names, not by reading its heading, and participants should see that.
def with_province(columns: list[str], where: int = 0,
                  heading: str = "Tỉnh/thành phố") -> list[str]:
    out = list(columns)
    out.insert(where if where >= 0 else len(out) + 1 + where, heading)
    return out


ALL = ["Xã/phường", "Dân số 15-49 tuổi", "Ước tính quần thể nguy cơ cao",
       "Số người được tiếp cận cộng đồng", "Số bao cao su đã phát",
       "Số bơm kim tiêm đã phát", "Số người được xét nghiệm HIV",
       "Số ca dương tính", "Tỷ lệ dương tính (%)",
       "Số người nhiễm HIV đang quản lý", "Số người đang điều trị ARV",
       "Tỷ lệ điều trị ARV (%)", "Số người đang dùng PrEP",
       "Nhóm ưu tiên can thiệp"]

OUTREACH = ["Xã/phường", "Ước tính quần thể nguy cơ cao",
            "Số người được tiếp cận cộng đồng", "Số bao cao su đã phát",
            "Số bơm kim tiêm đã phát", "Nhóm ưu tiên can thiệp"]

TESTING = ["Xã/phường", "Dân số 15-49 tuổi", "Số người được xét nghiệm HIV",
           "Số ca dương tính", "Tỷ lệ dương tính (%)"]

TREATMENT = ["Xã/phường", "Số người nhiễm HIV đang quản lý",
             "Số người đang điều trị ARV", "Tỷ lệ điều trị ARV (%)",
             "Số người đang dùng PrEP"]


def build(province: str, slug: str, rows: list[dict], index: int) -> Path:
    """One workbook, laid out unlike any of the other ten."""
    book = Workbook()
    book.remove(book.active)
    period = "6 tháng đầu năm 2026"
    long_title = (f"BÁO CÁO CHƯƠNG TRÌNH PHÒNG, CHỐNG HIV/AIDS - {province.upper()}")

    if index == 0:
        # the plainest possible: names on row 1, nothing else, no dictionary
        sheet_of(book, "Data", rows, with_province(ALL), header_row=1,
                 province=province)

    elif index == 1:
        columns = with_province(ALL, -1)          # province in the LAST column
        sheet_of(book, "Số liệu xã phường", rows, columns, header_row=4,
                 preamble=[(long_title, TITLE), (f"Kỳ báo cáo: {period}", SUB)],
                 province=province)
        guidance_sheet(book, province, columns)

    elif index == 2:
        sheet_of(book, "Tổng hợp", rows, with_province(ALL, 1, "Tỉnh"),
                 header_row=6, province=province, preamble=[
            ("SỞ Y TẾ " + province.upper(), SUB),
            ("TRUNG TÂM KIỂM SOÁT BỆNH TẬT", SUB),
            ("", SUB),
            (long_title, TITLE),
            (f"Kỳ báo cáo: {period}   -   Ngày lập: 20/08/2026", SUB),
        ])

    elif index == 3:
        outreach, testing = with_province(OUTREACH), with_province(TESTING)
        sheet_of(book, "Can thiệp cộng đồng", rows, outreach, header_row=3,
                 province=province,
                 preamble=[(f"Tiếp cận cộng đồng và cấp phát vật dụng - {province}",
                            TITLE), (f"Kỳ báo cáo: {period}", SUB)])
        sheet_of(book, "Xét nghiệm", rows, testing, header_row=3,
                 province=province, preamble=[
            (f"Xét nghiệm HIV - {province}", TITLE), (f"Kỳ báo cáo: {period}", SUB)])
        guidance_sheet(book, province, sorted(set(outreach + testing)))

    elif index == 4:
        sheet_of(book, "Bao cao", rows, with_province(ALL), header_row=9,
                 province=province, preamble=[
            ("CỘNG HÒA XÃ HỘI CHỦ NGHĨA VIỆT NAM", SUB),
            ("Độc lập - Tự do - Hạnh phúc", SUB),
            ("", SUB),
            ("SỞ Y TẾ " + province.upper(), SUB),
            ("TRUNG TÂM KIỂM SOÁT BỆNH TẬT", SUB),
            ("", SUB),
            (long_title, TITLE),
            (f"Kỳ báo cáo: {period}", SUB),
        ], footer="Người lập biểu: ......................")

    elif index == 5:
        columns = with_province(TREATMENT, 1)
        sheet_of(book, "Điều trị ARV và PrEP", rows, columns, header_row=5,
                 province=province,
                 preamble=[(f"Điều trị ARV và dự phòng PrEP - {province}", TITLE),
                           (f"Kỳ báo cáo: {period}", SUB), ("", SUB),
                           ("Đơn vị tính: người", NOTE)])
        guidance_sheet(book, province, columns)

    elif index == 6:
        sheet_of(book, "SO LIEU", rows, with_province(ALL, -1, "Tỉnh"),
                 header_row=2, preamble=[(f"{province} - {period}", TITLE)],
                 province=province)

    elif index == 7:
        columns = ["Tỉnh/thành phố"] + ALL
        sheet_of(book, "Tổng hợp toàn tỉnh", rows, columns, header_row=7,
                 preamble=[("SỞ Y TẾ " + province.upper(), SUB),
                           ("TRUNG TÂM KIỂM SOÁT BỆNH TẬT", SUB), ("", SUB),
                           (long_title, TITLE),
                           (f"Kỳ báo cáo: {period}", SUB), ("", SUB)],
                 province=province)
        guidance_sheet(book, province, columns)

    elif index == 8:
        sheet_of(book, "Số liệu", rows, with_province(ALL, 2), header_row=4,
            province=province, preamble=[
            (long_title, TITLE), (f"Kỳ báo cáo: {period}", SUB)],
            footer=("Ghi chú: xã không có chương trình bơm kim tiêm ghi 0. "
                    "Số liệu giả lập phục vụ tập huấn."))

    elif index == 9:
        sheet_of(book, "Chi tiet xa phuong", rows, with_province(ALL),
                 header_row=10, province=province, preamble=[
            ("CỘNG HÒA XÃ HỘI CHỦ NGHĨA VIỆT NAM", SUB),
            ("Độc lập - Tự do - Hạnh phúc", SUB),
            ("-----------------------", SUB),
            ("", SUB),
            ("SỞ Y TẾ " + province.upper(), SUB),
            ("TRUNG TÂM KIỂM SOÁT BỆNH TẬT", SUB),
            ("", SUB),
            (long_title, TITLE),
            (f"Kỳ báo cáo: {period}   -   Đơn vị lập: Khoa Phòng, chống HIV/AIDS",
             SUB),
        ])
        guidance_sheet(book, province, with_province(ALL))

    else:
        # two periods on two sheets, so the same file can be mapped twice
        half = max(1, len(rows) // 2)
        columns = with_province(ALL)
        sheet_of(book, "Quy I-2026", rows[:half] + rows[half:], columns,
                 header_row=3, province=province,
                 preamble=[(f"{province} - Quý I/2026", TITLE),
                           ("Đơn vị: Trung tâm Kiểm soát bệnh tật", SUB)])
        sheet_of(book, "Quy II-2026", rows, columns, header_row=3,
                 province=province,
                 preamble=[(f"{province} - Quý II/2026", TITLE),
                           ("Đơn vị: Trung tâm Kiểm soát bệnh tật", SUB)])

    path = OUT / f"{slug}_du_lieu_thuc_hanh.xlsx"
    book.save(path)
    return path


def main() -> int:
    from emap import dataio

    root = dataio.shapefile_root(ROOT)
    shp = next((root / "viet-nam" / "commune").glob("*.shp"))
    frame = gpd.read_file(shp, columns=["ten_tinh", "ten_xa"])
    if dataio._encoding_repair(gpd.read_file(shp, rows=5), shp) is not None:
        frame = gpd.read_file(shp, columns=["ten_tinh", "ten_xa"],
                              encoding="utf-8")

    OUT.mkdir(parents=True, exist_ok=True)
    print(f"{len(PROVINCES)} workbooks in {OUT}\n")
    print(f"  {'tệp':44} {'xã':>4}  {'dòng tiêu đề':>12}  từ điển")
    for index, (province, slug) in enumerate(PROVINCES):
        rows = communes(frame, province)
        if not rows:
            print(f"  {slug}: no communes found for {province}", file=sys.stderr)
            continue
        path = build(province, slug, rows, index)
        from openpyxl import load_workbook

        book = load_workbook(path)
        first = book[book.sheetnames[0]]
        header = next(r for r in range(1, 15)
                      if first.cell(row=r, column=1).value in
                      ("Xã/phường", "Tỉnh/thành phố", "Tỉnh"))
        print(f"  {path.name:44} {len(rows):4}  {header:12}  "
              f"{'có' if 'Từ điển dữ liệu' in book.sheetnames else '-'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
