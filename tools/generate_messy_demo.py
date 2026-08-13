"""A workbook built to be as awkward as a real one, and then some.

Every other fixture in ``input/`` is awkward in one way: ``bao_cao_so_y_te``
has a title block and merged cells, the PEPFAR export is long-format, the
commune file is sparse. This one puts them all in a single workbook, together
with the traps this project has actually been caught by, so a run against it
exercises the whole reading path at once.

The point is not that the skill draws a beautiful map from it. The point is what
it *says* on the way — which sheets it refuses, which columns it warns about, and
which place names it sends for review. Each trap below is numbered, and the
number is what to check the output against.

Nothing here is real. The figures are invented, the misspellings are deliberate,
and two of the place names do not exist.
"""

from __future__ import annotations

import random
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parents[1] / "input" / "du_lieu_lung_tung.xlsx"
SEED = 20260808

TITLE = Font(bold=True, size=13)
HEAD = Font(bold=True, size=10)
NOTE = Font(italic=True, size=9, color="6B7780")
FILL = PatternFill("solid", fgColor="EAF1F6")
CENTRE = Alignment(horizontal="center", vertical="center", wrap_text=True)


#: Place names as a data clerk really writes them. The comment on each says which
#: part of the matching path it is meant to exercise.
PLACES = [
    ("Hà Nội", "clean"),
    ("  Hải Phòng  ", "T14 leading and trailing spaces"),
    ("Đà  Nẵng", "T14 a double space inside the name"),
    ("AN GIANG", "T18 shouted"),
    ("tuyên quang", "T18 whispered"),
    ("Tỉnh Lào Cai", "T15 an administrative prefix"),
    ("TP. Cần Thơ", "T15 a city prefix"),
    ("Thừa Thiên Huế", "T12 the name from before the 2025 merger — now Huế"),
    ("Bà Rịa - Vũng Tàu", "T12 merged into TP. Hồ Chí Minh"),
    ("Hải Dương", "T12 merged into Hải Phòng"),
    ("Nghệ Ann", "T13 a typo — one letter too many"),
    ("Thanh Hoá", "T13 the other spelling of Thanh Hóa"),
    ("Thủ Đức", "T16 a district, a tier abolished in 2025"),
    ("Tỉnh Hoa Lư", "T17 no such province"),
    ("Quảng Ninh", "clean"),
    ("Lâm Đồng", "clean"),
    ("Cà Mau", "clean"),
    ("Sơn La", "clean"),
    ("Điện Biên", "clean"),
    ("Hà Nội", "T19 a second row for a province already listed"),
]


def sheet_summary(wb, rng: random.Random) -> None:
    """The main table: T1–T11, T20–T22.

    A title block, a header split over two rows with a merged span, provinces
    written once and blank underneath, and a set of column names chosen to
    collide with the skill's own keyword lists.
    """
    ws = wb.create_sheet("Tổng hợp 2026")

    ws["A1"] = "SỞ Y TẾ — BÁO CÁO TỔNG HỢP CHƯƠNG TRÌNH 2026"
    ws["A1"].font = TITLE
    ws["A2"] = "Đơn vị báo cáo: các tỉnh/thành phố trực thuộc trung ương"
    ws["A3"] = "Kỳ số 04/2026 · ban hành ngày 08/08/2026"
    ws["A4"] = "(SỐ LIỆU GIẢ LẬP — chỉ dùng để thử phần mềm, không dùng cho báo cáo)"
    ws["A4"].font = NOTE
    # row 5 left blank on purpose: T1, the header does not start at row 1

    # T2 two-tier header. "Số ca phát hiện" spans two columns named on row 7.
    top = ["Mã ĐV", "Tỉnh/thành phố", "Kỳ báo cáo", "Quy mô dân số",
           "Số ca phát hiện", None, "Số ca mắc trong ngày", "Thang điểm ưu tiên",
           "Tỷ lệ điều trị (%)", "Tỷ lệ bao phủ", "Ngân sách (triệu đồng)",
           "Ghi chú", "Miền Nam"]
    sub = [None, None, None, None, "Nam", "Nữ", None, None, None, None, None, None, None]
    for c, (a, b) in enumerate(zip(top, sub), start=1):
        if a is not None:
            ws.cell(row=6, column=c, value=a).font = HEAD
            ws.cell(row=6, column=c).fill = FILL
            ws.cell(row=6, column=c).alignment = CENTRE
        if b is not None:
            ws.cell(row=7, column=c, value=b).font = HEAD
            ws.cell(row=7, column=c).fill = FILL
            ws.cell(row=7, column=c).alignment = CENTRE
    ws.merge_cells("E6:F6")                       # T2 the span
    for col in (1, 2, 3, 4, 7, 8, 9, 10, 11, 12, 13):
        letter = get_column_letter(col)
        ws.merge_cells(f"{letter}6:{letter}7")    # T2 the single-tier headings

    row = 8
    for i, (place, _why) in enumerate(PLACES):
        nam = rng.randint(40, 190)
        nu = rng.randint(25, 130)
        pop = rng.randint(700_000, 8_500_000)

        ws.cell(row=row, column=1, value=f"DV-{i + 1:03d}")          # T21 identifier
        ws.cell(row=row, column=2, value=place)
        ws.cell(row=row, column=3, value="Năm 2026")
        # T7 a population written as text, the Vietnamese way
        ws.cell(row=row, column=4, value=f"{pop:,}".replace(",", ".") if i % 3 == 0 else pop)
        ws.cell(row=row, column=5, value=nam)                        # T4 "Nam" is male
        ws.cell(row=row, column=6, value=nu)
        ws.cell(row=row, column=7, value=rng.randint(1, 12))         # T6 quantity, not a period
        ws.cell(row=row, column=8, value=rng.randint(1, 5))          # T5 "Thang" is a scale
        # T8 one figure above 100, T9 one below zero — both are data errors the
        # guardrails are supposed to name rather than quietly draw
        pct = {2: 118.4, 5: -3.2}.get(i, round(rng.uniform(62, 97), 1))
        ws.cell(row=row, column=9,
                value=str(pct).replace(".", ",") if i % 4 == 0 else pct)   # T7 again
        ws.cell(row=row, column=10, value=round(rng.uniform(0.21, 0.94), 3))  # T10 a 0–1 scale
        ws.cell(row=row, column=11, value=rng.randint(800, 15_000))          # T20 money
        # column 12 is left empty for every row: T11
        ws.cell(row=row, column=13, value=rng.choice(["Có", "Không"]))       # T4 again
        row += 1

    # T3 provinces written once, blank underneath. Hà Nội appears twice (T19), so
    # the two rows are merged into one visual block the way a clerk would.
    ws.merge_cells(start_row=8, start_column=2, end_row=8, end_column=2)

    widths = [9, 20, 12, 15, 8, 8, 14, 13, 15, 12, 16, 10, 11]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w


def sheet_quarters(wb, rng: random.Random) -> None:
    """Long format, plus a second table underneath: T23–T26."""
    ws = wb.create_sheet("Diễn biến")

    ws["A1"] = "Bảng 1 — diễn biến theo kỳ"
    ws["A1"].font = TITLE
    head = ["Tỉnh/thành phố", "Kỳ", "Nhóm", "Số ca", "Tỷ lệ dương tính (%)"]
    for c, name in enumerate(head, start=1):
        ws.cell(row=2, column=c, value=name).font = HEAD
        ws.cell(row=2, column=c).fill = FILL

    # T24 three ways of writing a period, in one column
    periods = ["Quý I/2026", "2026-Q2", "Năm 2026"]
    groups = ["Nam", "Nữ", "Tổng"]          # T25 a total row beside its own parts
    row = 3
    for place in ["Hà Nội", "Hải Phòng", "Nghệ An", "Cà Mau"]:
        for period in periods:
            for group in groups:            # T23 one unit over many rows
                ws.cell(row=row, column=1, value=place)
                ws.cell(row=row, column=2, value=period)
                ws.cell(row=row, column=3, value=group)
                ws.cell(row=row, column=4, value=rng.randint(20, 240))
                ws.cell(row=row, column=5, value=round(rng.uniform(0.4, 2.6), 2))
                row += 1

    # T26 a second, unrelated table in the same sheet
    row += 2
    ws.cell(row=row, column=1, value="Bảng 2 — chỉ tiêu giao").font = TITLE
    row += 1
    for c, name in enumerate(["Tỉnh/thành phố", "Chỉ tiêu 2026", "Đã đạt"], start=1):
        ws.cell(row=row, column=c, value=name).font = HEAD
        ws.cell(row=row, column=c).fill = FILL
    row += 1
    for place in ["Hà Nội", "Hải Phòng", "Nghệ An"]:
        ws.cell(row=row, column=1, value=place)
        ws.cell(row=row, column=2, value=rng.randint(500, 2000))
        ws.cell(row=row, column=3, value=rng.randint(300, 1900))
        row += 1

    for c, w in enumerate([20, 14, 10, 10, 20], start=1):
        ws.column_dimensions[get_column_letter(c)].width = w


def sheet_points(wb, rng: random.Random) -> None:
    """Coordinates, two of them the wrong way round: T27–T28."""
    ws = wb.create_sheet("Cơ sở y tế")
    head = ["Tên cơ sở", "Tỉnh/thành phố", "Loại hình", "Kinh độ", "Vĩ độ",
            "Số lượt khám"]
    for c, name in enumerate(head, start=1):
        ws.cell(row=1, column=c, value=name).font = HEAD
        ws.cell(row=1, column=c).fill = FILL

    places = ["Hà Nội", "Hải Phòng", "Đà Nẵng", "Cần Thơ", "Nghệ An", "Lâm Đồng"]
    for i, place in enumerate(places, start=2):
        lon = round(rng.uniform(103.5, 109.2), 4)
        lat = round(rng.uniform(8.6, 23.2), 4)
        if i == 4:                          # T27 longitude and latitude swapped
            lon, lat = lat, lon
        ws.cell(row=i, column=1, value=f"Trung tâm y tế {place}")
        ws.cell(row=i, column=2, value=place)
        ws.cell(row=i, column=3, value=rng.choice(["Công lập", "Tư nhân"]))  # T28
        ws.cell(row=i, column=4, value=lon)
        ws.cell(row=i, column=5, value=lat)
        ws.cell(row=i, column=6, value=rng.randint(400, 9000))

    for c, w in enumerate([26, 18, 12, 11, 11, 14], start=1):
        ws.column_dimensions[get_column_letter(c)].width = w


def sheet_pivot(wb) -> None:
    """A pivot table: the shape that has no column names to speak of."""
    ws = wb.create_sheet("Pivot")
    ws["A1"] = "Tổng hợp theo miền"
    ws["A1"].font = TITLE
    ws["A3"] = "Miền"
    ws["B3"] = "Tổng"
    for r, (name, value) in enumerate(
            [("Bắc", 4821), ("Trung", 2109), ("Nam", 6640)], start=4):
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=value)
    # trailing unnamed columns, which is what a copied pivot leaves behind
    for c in range(3, 8):
        ws.cell(row=3, column=c, value=None)
        ws.cell(row=4, column=c, value=rng_blank())


def rng_blank():
    return None


def sheet_notes(wb) -> None:
    """Prose. Nothing to map, and the skill should say so rather than try."""
    ws = wb.create_sheet("Hướng dẫn")
    lines = [
        "HƯỚNG DẪN GHI BIỂU",
        "",
        "1. Cột 'Tỉnh/thành phố' ghi theo danh mục hành chính hiện hành.",
        "2. Cột tỷ lệ ghi theo phần trăm, một chữ số thập phân.",
        "3. Số liệu trong tệp này là GIẢ LẬP, dùng để thử phần mềm.",
        "",
        "Liên hệ: phòng kế hoạch tổng hợp.",
    ]
    for r, line in enumerate(lines, start=1):
        ws.cell(row=r, column=1, value=line)
    ws.column_dimensions["A"].width = 70


def main() -> None:
    rng = random.Random(SEED)
    wb = Workbook()
    wb.remove(wb.active)

    sheet_summary(wb, rng)
    sheet_quarters(wb, rng)
    sheet_points(wb, rng)
    sheet_pivot(wb)
    sheet_notes(wb)
    wb.create_sheet("Sheet1")               # an empty sheet, as always survives

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT}  ({OUT.stat().st_size / 1024:.0f} KB, "
          f"{len(wb.sheetnames)} sheets: {', '.join(wb.sheetnames)})")


if __name__ == "__main__":
    main()
