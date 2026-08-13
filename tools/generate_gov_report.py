"""Build a workbook shaped like a real Vietnamese agency report.

Three habits that every previous fixture was too tidy to have, and that between
them break a straightforward read:

* a title block above the table, so the column names are not on row 1;
* a **two-tier header**, where "Số ca phát hiện" spans two columns that are
  named "Nam" and "Nữ" on the row below;
* **merged cells down the province column**, so a province is written once and
  the rows beneath it are blank — blank meaning "same as above", not "missing".

Sheet two puts two separate tables in one sheet, which is what happens when
somebody appends a second table rather than opening a new sheet.

Everything here is invented. It is shaped like real reporting and contains no
real figures.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parents[1] / "input" / "bao_cao_so_y_te.xlsx"

PROVINCES = [
    ("Hà Nội", [("Quý I", 128, 94, 91.4), ("Quý II", 141, 88, 92.6)]),
    ("Nghệ An", [("Quý I", 76, 52, 84.2), ("Quý II", 81, 49, 85.9)]),
    ("Đà Nẵng", [("Quý I", 44, 31, 93.7), ("Quý II", 47, 35, 94.1)]),
    ("Cần Thơ", [("Quý I", 39, 28, 88.0), ("Quý II", 42, 30, 88.8)]),
    ("Khánh Hòa", [("Quý I", 55, 37, 90.3), ("Quý II", 58, 41, 90.9)]),
]

BUDGET = [("Hà Nội", 4_820_000_000), ("Nghệ An", 2_150_000_000),
          ("Đà Nẵng", 1_640_000_000), ("Cần Thơ", 1_280_000_000),
          ("Khánh Hòa", 1_710_000_000)]


def _title(ws, row: int, text: str, span: int, *, size: float = 12.0) -> None:
    ws.cell(row=row, column=1, value=text).font = Font(bold=True, size=size)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")


def build_report_sheet(ws) -> None:
    _title(ws, 1, "BÁO CÁO TỔNG HỢP CHƯƠNG TRÌNH HIV — NĂM 2026", 6)
    ws.cell(row=2, column=1, value="Đơn vị báo cáo: Sở Y tế các tỉnh/thành phố")
    ws.cell(row=3, column=1, value="(Số liệu giả lập, chỉ dùng để thử phần mềm)")
    # row 4 is left empty, as report templates usually do

    header_top, header_bottom = 5, 6
    ws.cell(row=header_top, column=1, value="TT")
    ws.cell(row=header_top, column=2, value="Tỉnh/thành phố")
    ws.cell(row=header_top, column=3, value="Kỳ báo cáo")
    ws.cell(row=header_top, column=4, value="Số ca phát hiện")
    ws.cell(row=header_bottom, column=4, value="Nam")
    ws.cell(row=header_bottom, column=5, value="Nữ")
    ws.cell(row=header_top, column=6, value="Tỷ lệ điều trị (%)")

    # the single-tier columns span both header rows; the count column spans two
    # columns on the top row only
    for column in (1, 2, 3, 6):
        ws.merge_cells(start_row=header_top, start_column=column,
                       end_row=header_bottom, end_column=column)
    ws.merge_cells(start_row=header_top, start_column=4,
                   end_row=header_top, end_column=5)
    for row in (header_top, header_bottom):
        for column in range(1, 7):
            cell = ws.cell(row=row, column=column)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)

    row = header_bottom + 1
    for index, (province, periods) in enumerate(PROVINCES, start=1):
        first = row
        for period, male, female, rate in periods:
            ws.cell(row=row, column=3, value=period)
            ws.cell(row=row, column=4, value=male)
            ws.cell(row=row, column=5, value=female)
            ws.cell(row=row, column=6, value=rate)
            row += 1
        ws.cell(row=first, column=1, value=index)
        ws.cell(row=first, column=2, value=province)
        # the province is written once and merged down over its own rows
        ws.merge_cells(start_row=first, start_column=1, end_row=row - 1, end_column=1)
        ws.merge_cells(start_row=first, start_column=2, end_row=row - 1, end_column=2)
        for column in (1, 2):
            ws.cell(row=first, column=column).alignment = Alignment(
                horizontal="center", vertical="center")

    for column, width in enumerate((6, 20, 14, 10, 10, 16), start=1):
        ws.column_dimensions[get_column_letter(column)].width = width


def build_two_tables_sheet(ws) -> None:
    _title(ws, 1, "Bảng 1. Số ca phát hiện theo tỉnh", 2, size=11.0)
    ws.cell(row=2, column=1, value="Tỉnh/thành phố")
    ws.cell(row=2, column=2, value="Số ca phát hiện")
    row = 3
    for province, periods in PROVINCES:
        ws.cell(row=row, column=1, value=province)
        ws.cell(row=row, column=2, value=sum(p[1] + p[2] for p in periods))
        row += 1

    row += 3                                   # three blank rows between tables
    _title(ws, row, "Bảng 2. Ngân sách chương trình theo tỉnh", 2, size=11.0)
    ws.cell(row=row + 1, column=1, value="Tỉnh/thành phố")
    ws.cell(row=row + 1, column=2, value="Ngân sách (VND)")
    row += 2
    for province, amount in BUDGET:
        ws.cell(row=row, column=1, value=province)
        ws.cell(row=row, column=2, value=amount)
        row += 1

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 22


def build_plain_sheet(ws) -> None:
    """A control: no merges, no title block, header on row 1."""
    ws.append(["Tỉnh/thành phố", "Số ca phát hiện", "Tỷ lệ điều trị (%)"])
    for province, periods in PROVINCES:
        ws.append([province, sum(p[1] + p[2] for p in periods),
                   round(sum(p[3] for p in periods) / len(periods), 1)])
    ws.column_dimensions["A"].width = 20


def main() -> None:
    book = Workbook()
    build_report_sheet(book.active)
    book.active.title = "Tổng hợp năm 2026"
    build_two_tables_sheet(book.create_sheet("Hai bảng"))
    build_plain_sheet(book.create_sheet("Bảng phẳng"))
    OUT.parent.mkdir(parents=True, exist_ok=True)
    book.save(OUT)
    print(f"đã ghi {OUT}")


if __name__ == "__main__":
    main()
