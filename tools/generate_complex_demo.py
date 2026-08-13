from __future__ import annotations

import argparse
import random
from pathlib import Path

import geopandas as gpd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_BLUE = "0056A7"
HEADER_TEAL = "007C91"
PALE_YELLOW = "FFF4CE"
WHITE = "FFFFFF"


def style_table(ws, freeze: str = "A2") -> None:
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=HEADER_BLUE)
        cell.font = Font(name="Arial", color=WHITE, bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 34
    for column in ws.columns:
        values = [str(cell.value or "") for cell in column]
        width = min(max(max(map(len, values)) + 2, 12), 34)
        ws.column_dimensions[get_column_letter(column[0].column)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def select_communes(frame, province: str, count: int = 12):
    subset = frame.loc[frame["ten_tinh"] == province].copy()
    subset = subset.sort_values(["loai", "stt", "ten_xa"])
    if len(subset) < count:
        raise ValueError(f"Not enough communes for {province}: {len(subset)}")
    step = max(1, len(subset) // count)
    selected = subset.iloc[::step].head(count).copy()
    return selected.sort_values("ten_xa")


def messy_name(name: str, province: str, index: int) -> tuple[str, str]:
    if province == "Hà Nội" and index == 2:
        replacements = str.maketrans(
            "àáạảãâầấậẩẫăằắặẳẵèéẹẻẽêềếệểễìíịỉĩòóọỏõôồốộổỗơờớợởỡùúụủũưừứựửữỳýỵỷỹđ",
            "aaaaaaaaaaaaaaaaaeeeeeeeeeeeiiiiiooooooooooooooooouuuuuuuuuuuyyyyyd",
        )
        return f"Xa {name.lower().translate(replacements).title()}", "Tên xã không dấu và có tiền tố rút gọn"
    if province == "Huế" and index == 5:
        return f"P. {name}", "Tên phường dùng tiền tố viết tắt"
    if province == "Cần Thơ" and index == 8 and len(name) > 4:
        return name[:-1], "Tên địa danh thiếu ký tự cuối"
    return name, ""


def build_rows(communes) -> tuple[list[list[object]], list[list[object]]]:
    rng = random.Random(20260805)
    main_rows: list[list[object]] = []
    monthly_rows: list[list[object]] = []
    provinces = ["Hà Nội", "Huế", "Cần Thơ"]
    partners = ["Đơn vị A", "Đơn vị B", "Đơn vị C"]

    for province_index, province in enumerate(provinces):
        selected = select_communes(communes, province)
        for index, record in enumerate(selected.itertuples(index=False)):
            name, quality_note = messy_name(record.ten_xa, province, index)
            population = int(record.dan_so)
            target_population = max(1200, int(round(population * (0.075 + 0.01 * (index % 3)))))
            coverage_2025 = round(48 + province_index * 7 + index * 2.7 + rng.uniform(-4, 4), 1)
            coverage_2025 = min(91.0, max(35.0, coverage_2025))
            delta = round(rng.uniform(-7, 16), 1)
            coverage_2026 = min(98.5, max(30.0, round(coverage_2025 + delta, 1)))
            reached_2026 = int(round(target_population * coverage_2026 / 100))
            tested_2026 = max(120, int(round(reached_2026 * rng.uniform(0.18, 0.42))))
            positivity = max(0.4, rng.uniform(1.0, 9.5) + (2 if index in {1, 9} else 0))
            cases_2026 = int(round(tested_2026 * positivity / 100))
            cases_2025 = max(0, int(round(cases_2026 * rng.uniform(0.55, 1.45))))
            incidence = round(cases_2026 / population * 100_000, 1)
            change_pp = round(coverage_2026 - coverage_2025, 1)
            status = (
                "Cần tăng cường"
                if coverage_2026 < 60 or positivity >= 8
                else "Theo dõi sát"
                if coverage_2026 < 75 or positivity >= 5
                else "Đạt tiến độ"
            )
            priority = "Rất cao" if positivity >= 8 else "Cao" if positivity >= 5 else "Thường quy"
            budget = int(round((180_000_000 + population * rng.uniform(650, 1300)) / 1_000_000) * 1_000_000)
            completeness = 100 if index % 7 else 92
            note = quality_note
            if completeness < 100:
                note = "; ".join(filter(None, [note, "Thiếu báo cáo của một điểm dịch vụ"]))
            if province == "Huế" and index == 9:
                budget = None
                note = "; ".join(filter(None, [note, "Chưa cập nhật ngân sách quý"]))
            if province == "Cần Thơ" and index == 10:
                coverage_2025 = None
                change_pp = None
                note = "; ".join(filter(None, [note, "Thiếu số liệu nền năm 2025"]))

            main_rows.append(
                [
                    province,
                    name,
                    "Quý I/2026",
                    partners[(index + province_index) % len(partners)],
                    population,
                    target_population,
                    reached_2026,
                    tested_2026,
                    cases_2025,
                    cases_2026,
                    incidence,
                    round(cases_2026 / tested_2026 * 100, 1),
                    coverage_2025,
                    coverage_2026,
                    change_pp,
                    priority,
                    status,
                    budget,
                    completeness,
                    note,
                ]
            )

            monthly_cases = []
            for month in range(10, 13):
                monthly_cases.append(max(0, int(round(cases_2025 / 3 + rng.uniform(-2, 2)))))
                monthly_rows.append(
                    [province, name, f"2025-{month:02d}", monthly_cases[-1], partners[(index + month) % 3]]
                )
            for month in range(1, 4):
                monthly_cases.append(max(0, int(round(cases_2026 / 3 + rng.uniform(-2, 2)))))
                monthly_rows.append(
                    [province, name, f"2026-{month:02d}", monthly_cases[-1], partners[(index + month) % 3]]
                )

    return main_rows, monthly_rows


def create_workbook(shapefile: Path, output: Path) -> None:
    communes = gpd.read_file(shapefile, ignore_geometry=True)
    main_rows, monthly_rows = build_rows(communes)

    workbook = Workbook()
    ws = workbook.active
    ws.title = "Dữ liệu xã hiện tại"
    headers = [
        "Tỉnh/thành phố",
        "Xã/phường",
        "Kỳ báo cáo",
        "Đơn vị hỗ trợ",
        "Dân số",
        "Dân số mục tiêu",
        "Số người tiếp cận 2026",
        "Số người xét nghiệm 2026",
        "Số ca phát hiện 2025",
        "Số ca phát hiện 2026",
        "Tỷ suất ca phát hiện/100.000 dân",
        "Tỷ lệ dương tính (%)",
        "Bao phủ 2025 (%)",
        "Bao phủ 2026 (%)",
        "Thay đổi bao phủ (điểm %)",
        "Mức ưu tiên",
        "Trạng thái can thiệp",
        "Ngân sách 2026 (VND)",
        "Mức hoàn thiện báo cáo (%)",
        "Ghi chú chất lượng dữ liệu",
    ]
    ws.append(headers)
    for row in main_rows:
        ws.append(row)
    style_table(ws)
    for row in range(2, ws.max_row + 1):
        ws.cell(row, 18).number_format = '#,##0 "₫"'
        for col in (11, 12, 13, 14, 15, 19):
            ws.cell(row, col).number_format = "0.0"
        if ws.cell(row, 20).value:
            ws.cell(row, 20).fill = PatternFill("solid", fgColor=PALE_YELLOW)

    detail = workbook.create_sheet("Diễn biến theo tháng")
    detail.append(["Tỉnh/thành phố", "Xã/phường", "Tháng", "Số ca phát hiện", "Đơn vị báo cáo"])
    for row in monthly_rows:
        detail.append(row)
    style_table(detail)

    dictionary = workbook.create_sheet("Từ điển dữ liệu")
    dictionary.append(["Tên cột", "Ý nghĩa", "Kiểu dữ liệu", "Đơn vị", "Cách tổng hợp phù hợp"])
    definitions = [
        ("Tỉnh/thành phố", "Tên đơn vị hành chính cấp tỉnh sau sắp xếp", "Địa lý", "Tên", "Không tổng hợp"),
        ("Xã/phường", "Tên đơn vị hành chính cấp xã; không có mã hành chính", "Địa lý", "Tên", "Ghép trong phạm vi tỉnh"),
        ("Số ca phát hiện 2025/2026", "Số ca mới phát hiện trong kỳ tương ứng", "Số nguyên", "Ca", "Cộng"),
        ("Tỷ suất ca phát hiện/100.000 dân", "Số ca năm 2026 chia dân số, nhân 100.000", "Số thập phân", "Ca/100.000 dân", "Tính lại từ tử số và mẫu số"),
        ("Tỷ lệ dương tính (%)", "Số ca 2026 chia số người xét nghiệm 2026", "Số thập phân", "%", "Tính lại có trọng số"),
        ("Bao phủ 2025/2026 (%)", "Số người tiếp cận chia dân số mục tiêu", "Số thập phân", "%", "Tính lại có trọng số"),
        ("Thay đổi bao phủ (điểm %)", "Bao phủ 2026 trừ bao phủ 2025", "Số thập phân có dấu", "Điểm phần trăm", "Không cộng"),
        ("Mức ưu tiên", "Nhóm ưu tiên chương trình tổng hợp", "Phân loại có thứ tự", "Nhóm", "Không tổng hợp"),
        ("Trạng thái can thiệp", "Trạng thái quản lý dựa trên bao phủ và tỷ lệ dương tính", "Phân loại", "Nhóm", "Không tổng hợp"),
        ("Ngân sách 2026 (VND)", "Ngân sách kế hoạch năm 2026", "Tiền tệ", "VND", "Cộng"),
        ("Mức hoàn thiện báo cáo (%)", "Tỷ lệ điểm dịch vụ đã gửi báo cáo", "Số thập phân", "%", "Bình quân có trọng số nếu cần"),
    ]
    for row in definitions:
        dictionary.append(row)
    style_table(dictionary)

    readme = workbook.create_sheet("Hướng dẫn")
    readme.sheet_view.showGridLines = False
    readme.column_dimensions["A"].width = 26
    readme.column_dimensions["B"].width = 95
    rows = [
        ("Bộ dữ liệu", "Dữ liệu giám sát chương trình cấp xã giả lập, dùng để kiểm thử skill easy-map."),
        ("Phạm vi", "36 xã/phường thuộc Hà Nội, Huế và Cần Thơ; 12 đơn vị mỗi tỉnh."),
        ("Mục tiêu thử nghiệm", "Kiểm tra nhận diện cột địa lý, số đếm nguyên, tỷ lệ, thay đổi theo thời gian, phân loại và tiền tệ."),
        ("Chất lượng địa danh", "Ba tên xã/phường được viết không chuẩn có chủ ý và phải được xác nhận trước khi lập bản đồ."),
        ("Lưu ý", "Tất cả số liệu chương trình là giả lập. Ranh giới lấy từ shapefile đi kèm dự án."),
    ]
    for row in rows:
        readme.append(row)
    for cell in readme[1]:
        cell.fill = PatternFill("solid", fgColor=HEADER_TEAL)
        cell.font = Font(name="Arial", color=WHITE, bold=True)
    for row in readme.iter_rows():
        for cell in row:
            cell.font = Font(name="Arial", size=11, bold=cell.column == 1)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    readme.row_dimensions[1].height = 42

    workbook.calculation.fullCalcOnLoad = True
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    print(f"Created {output} with {len(main_rows)} main rows and {len(monthly_rows)} monthly rows")


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate a complex synthetic workbook for easy-map testing.")
    parser.add_argument("--shapefile", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    create_workbook(args.shapefile, args.output)


if __name__ == "__main__":
    main()
